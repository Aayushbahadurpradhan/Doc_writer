"""
generate_docs.py — Step 2 of the frontend pipeline.

Generates frontend documentation organized by route group.

Output structure:
  docs/frontend/
    index.md                         <- master index
    admin/
      README.md                      <- /admin group overview
      admin_dashboard.md             <- /admin
      admin_users.md                 <- /admin/users
    bill/
      README.md
      bill_list.md
      bill_detail.md
    undocumented/
      README.md                      <- components with no route or file
      missing_apis.md                <- APIs called by frontend pages

Each page .md includes:
  - Route / component / source file
  - Layout
  - Child components (imported locally)
  - Composables used
  - API dependencies (endpoint + method + caller + route)
  - State management
  - Unknowns / warnings
"""

import json
import os
import re
import sys
import time
from collections import defaultdict
from typing import Dict, List, Tuple

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from prompts.frontend_prompts import (pages_md_prompt, pages_md_system,
                                      resolve_dynamic_endpoint_prompt,
                                      undocumented_api_prompt)
from shared.ai_client import AIConfig, call_ai


def _extract_page_group(page: dict) -> str:
    """
    Determine the documentation folder group for a page.

    - Pages with path=UNKNOWN, inferred route, or missing component file → 'undocumented'
    - /admin/*   → 'admin'
    - /bill/list → 'bill'
    - /          → 'home'
    - /:id       → 'detail'
    """
    path = page.get("path", "UNKNOWN")
    has_no_file = any("Component file not found" in u for u in page.get("unknowns", []))
    inferred    = page.get("inferred", False)
    if path == "UNKNOWN" or has_no_file or inferred:
        return "undocumented"

    if not path or path == "/":
        return "home"

    parts = [p for p in path.strip("/").split("/") if p and not p.startswith(":")]
    if not parts:
        return "detail"

    group = parts[0].lower()
    # Keep plural as-is (admin, bills, agents all make sense as groups)
    return group or "home"


def _safe_page_filename(path: str) -> str:
    """Convert page path to safe filename."""
    safe = path.strip("/").replace("/", "_").replace("-", "_") or "index"
    safe = re.sub(r"[^\w]", "_", safe).strip("_") or "index"
    # Remove {id} type params from filename
    safe = re.sub(r"^\{.*?\}", "detail", safe)
    safe = re.sub(r"_\{.*?\}", "", safe)
    return safe + ".md"


def generate_pages_md(
    pages: List[dict],
    output_root: str,
    config: AIConfig,
    no_ai: bool = False,
) -> None:
    """
    Generate frontend docs organized by route group.
    Pages without a route or whose component file is missing go to undocumented/.
    Also writes frontend_detail.xlsx with one row per (page, API call).
    """
    output_root = os.path.abspath(output_root)
    os.makedirs(output_root, exist_ok=True)

    # ── Group pages ───────────────────────────────────────────────────────────
    groups: Dict[str, List[dict]] = defaultdict(list)
    all_api_calls: List[dict] = []
    all_excel_rows: List[dict] = []

    for page in pages:
        group = _extract_page_group(page)
        groups[group].append(page)

        for api_call in page.get("api_calls", []):
            endpoint = api_call.get("endpoint", "")
            if not endpoint:
                continue
            all_api_calls.append({
                "page_path":      page.get("path", "UNKNOWN"),
                "page_component": page.get("component", "UNKNOWN"),
                "endpoint":       endpoint,
                "method":         api_call.get("method", "UNKNOWN"),
                "via":            api_call.get("via", "direct"),
                "composable":     api_call.get("composable"),
                "called_from":    api_call.get("called_from", "UNKNOWN"),
                "dynamic":        api_call.get("dynamic", False),
                "vuex_action":    api_call.get("vuex_action"),
                "resolved_via":   api_call.get("resolved_via"),
                "env_config":     page.get("env_config") or {},
                "url_constants":  page.get("url_constants") or {},
            })

    # ── Generate per-group documentation ─────────────────────────────────────
    sys_msg = pages_md_system()

    for group_name in sorted(groups.keys()):
        group_pages = groups[group_name]
        group_dir   = os.path.normpath(os.path.join(output_root, group_name))
        os.makedirs(group_dir, exist_ok=True)

        _write_group_readme(group_name, group_pages, group_dir)

        for i, page in enumerate(group_pages):
            path       = page.get("path", "UNKNOWN")
            page_file  = _safe_page_filename(path)
            page_path  = os.path.join(group_dir, page_file)

            print(f"  [PAGE {i+1}/{len(group_pages)}] {group_name}/{page_file}")

            if no_ai or not config.use_ai:
                content = _skeleton_page(page)
                all_excel_rows.extend(_build_static_excel_rows(page))
            else:
                prompt = pages_md_prompt(page)
                try:
                    content = call_ai(prompt, config, system=sys_msg, max_tokens=2000)
                except Exception as e:
                    content = "[AI failed: {}]".format(str(e)[:80])
                if not content or content.startswith("[AI failed"):
                    print(f"     [WARN] AI failed - using skeleton")
                    content = _skeleton_page(page)
                    all_excel_rows.extend(_build_static_excel_rows(page))
                else:
                    excel_rows = _parse_excel_data(content)
                    if excel_rows:
                        # AI filled placeholders — merge with static fields
                        _merge_static_fields(page, excel_rows)
                        all_excel_rows.extend(excel_rows)
                    else:
                        all_excel_rows.extend(_build_static_excel_rows(page))
                if config.delay > 0:
                    time.sleep(config.delay)

            # Strip the EXCEL_DATA comment block before writing markdown
            md_content = re.sub(
                r"\n*<!--\s*EXCEL_DATA[\s\S]*?-->", "", content
            ).rstrip() + "\n"
            os.makedirs(os.path.dirname(page_path), exist_ok=True)
            with open(page_path, "w", encoding="utf-8") as f:
                f.write(md_content)

    # ── Master index ──────────────────────────────────────────────────────────
    _write_frontend_index(groups, output_root)

    # ── API network map (new) ─────────────────────────────────────────────────
    _write_api_network_map(groups, output_root)

    # ── Undocumented APIs ─────────────────────────────────────────────────────
    _write_missing_apis(all_api_calls, output_root, config, no_ai)

    # ── Excel output ──────────────────────────────────────────────────────────
    excel_path = _write_frontend_excel(all_excel_rows, output_root)
    if excel_path:
        print(f"  [EXCEL] Frontend API map -> {excel_path}")

    documented   = sum(len(v) for k, v in groups.items() if k != "undocumented")
    undocumented = len(groups.get("undocumented", []))
    print(f"\n  [OK] Frontend docs -> {output_root}")
    print(f"     Groups   : {', '.join(k for k in sorted(groups.keys()) if k != 'undocumented')}")
    print(f"     Documented   : {documented} pages")
    if undocumented:
        print(f"     Undocumented : {undocumented} components (no confirmed route)")


def _parse_excel_data(content: str) -> List[dict]:
    """Extract the EXCEL_DATA JSON array embedded in an AI markdown response."""
    m = re.search(r"<!--\s*EXCEL_DATA\s*\n([\s\S]*?)\n-->", content)
    if not m:
        return []
    try:
        rows = json.loads(m.group(1))
        return rows if isinstance(rows, list) else []
    except Exception:
        return []


def _merge_static_fields(page: dict, excel_rows: List[dict]) -> None:
    """Overwrite static fields (route, component_path, frontend_url) from page dict in-place."""
    path        = page.get("path", "")
    comp_file   = page.get("file") or page.get("component_file") or ""
    example_url = page.get("example_url") or ""
    for row in excel_rows:
        if not row.get("route"):
            row["route"] = path
        if not row.get("component_path"):
            row["component_path"] = comp_file
        if not row.get("frontend_url"):
            row["frontend_url"] = example_url


def _build_static_excel_rows(page: dict) -> List[dict]:
    """Build Excel rows from static page data — used when AI is disabled or fails."""
    path      = page.get("path", "UNKNOWN")
    component = page.get("component", "UNKNOWN")
    comp_file = page.get("file") or page.get("component_file") or ""
    api_calls = page.get("api_calls", [])
    example_url = page.get("example_url") or ""
    screen_name = (
        component
        .replace(".vue", "").replace(".jsx", "").replace(".tsx", "")
        .replace("_", " ").replace("-", " ")
    )

    rows = []
    for call in (api_calls or [{}]):
        trigger_name = call.get("trigger_name", "")
        trigger      = call.get("trigger", "")
        purpose      = call.get("purpose", "")

        if trigger_name:
            when = trigger_name
        elif trigger == "lifecycle":
            when = "On page load"
        elif trigger == "event_handler":
            when = "On user action"
        elif trigger == "watcher":
            when = "On data change"
        else:
            when = ""

        rows.append({
            "screen_name":    screen_name,
            "route":          path,
            "component_path": comp_file,
            "frontend_url":   example_url,
            "api_endpoint":   call.get("endpoint", ""),
            "http_method":    call.get("method", ""),
            "when_called":    when,
            "purpose":        purpose if purpose and purpose != "inline call" else "",
        })
    return rows


def _write_frontend_excel(rows: List[dict], output_root: str) -> str:
    """
    Write frontend_detail.xlsx with one row per (page, API call).

    Columns:
      #, Screen Name, Route / URL, Component Path, Frontend URL,
      API Endpoint, HTTP Method, When Called, Purpose
    """
    if not rows:
        return ""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("  [WARN] openpyxl not installed — skipping Excel output (pip install openpyxl)")
        return ""

    HEADERS = [
        "#",
        "Screen Name",
        "Route / URL",
        "Component Path",
        "Frontend URL",
        "API Endpoint",
        "HTTP Method",
        "When Called",
        "Purpose / What It Does",
    ]
    COL_WIDTHS = [5, 28, 25, 42, 45, 48, 13, 30, 55]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Frontend API Map"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap_top    = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = wrap_top

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width

    for row_num, row in enumerate(rows, 1):
        ws.append([
            row_num,
            row.get("screen_name", ""),
            row.get("route", ""),
            row.get("component_path", ""),
            row.get("frontend_url", ""),
            row.get("api_endpoint", ""),
            row.get("http_method", ""),
            row.get("when_called", ""),
            row.get("purpose", ""),
        ])
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row_num + 1, column=col_idx).alignment = wrap_top

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(HEADERS)).column_letter}1"

    excel_path = os.path.join(output_root, "frontend_api_map.xlsx")
    wb.save(excel_path)
    return excel_path


def _write_group_readme(group: str, pages: List[dict], group_dir: str) -> None:
    """Write README.md for each page group — network-trace focused."""
    readme_path = os.path.join(group_dir, "README.md")
    is_undoc    = group == "undocumented"

    with open(readme_path, "w", encoding="utf-8") as f:
        if is_undoc:
            f.write("# Undocumented Components\n\n")
            f.write(
                "These components were detected but either have no confirmed route mapping, "
                "their route was inferred from file path, or their source file could not be located.\n\n"
            )
        else:
            f.write(f"# /{group} Pages\n\n")
            f.write(f"Route prefix: **`/{group}`**\n\n")

        f.write("## Summary\n\n")
        f.write("| Route | Component | Frontend URL | # APIs | APIs Called |\n")
        f.write("|-------|-----------|-------------|--------|-------------|\n")

        for page in pages:
            path        = page.get("path", "UNKNOWN")
            component   = page.get("component", "UNKNOWN")
            api_calls   = page.get("api_calls", [])
            example_url = page.get("example_url") or "—"
            inferred    = page.get("inferred", False)

            page_file = _safe_page_filename(path)
            path_label = f"{path} _(inferred)_" if inferred else path

            # Compact unique endpoint list (max 4 shown inline)
            unique_eps = list(dict.fromkeys(
                c.get("endpoint", "")
                for c in api_calls
                if c.get("endpoint") and c.get("endpoint") != "UNKNOWN"
            ))
            if unique_eps:
                ep_str = ", ".join(f"`{e}`" for e in unique_eps[:4])
                if len(unique_eps) > 4:
                    ep_str += f" _(+{len(unique_eps)-4} more)_"
            else:
                ep_str = "—"

            f.write(
                f"| [{path_label}]({page_file}) "
                f"| `{component}` "
                f"| `{example_url}` "
                f"| {len(api_calls)} "
                f"| {ep_str} |\n"
            )

        f.write("\n---\n")


def _write_frontend_index(groups: Dict[str, List[dict]], output_root: str) -> None:
    """Write master index for all frontend pages."""
    index_path = os.path.join(output_root, "index.md")

    doc_groups  = {k: v for k, v in groups.items() if k != "undocumented"}
    undoc_pages = groups.get("undocumented", [])

    total_pages = sum(len(v) for v in doc_groups.values())
    total_apis  = sum(
        len(p.get("api_calls", [])) for pages in doc_groups.values() for p in pages
    )

    with open(index_path, "w", encoding="utf-8") as f:
        f.write("# Frontend Documentation\n\n")
        f.write(f"**Documented pages**: {total_pages} | **API calls detected**: {total_apis}")
        if undoc_pages:
            f.write(f" | **Undocumented**: {len(undoc_pages)}")
        f.write("\n\n")
        f.write("> See [api_network_map.md](api_network_map.md) for the full "
                "Frontend → Backend API dependency map.\n\n")

        f.write("## Page Groups\n\n")
        f.write("| Group | Route Prefix | Pages | APIs |\n")
        f.write("|-------|-------------|-------|------|\n")

        for group_name in sorted(doc_groups.keys()):
            gp      = doc_groups[group_name]
            api_cnt = sum(len(p.get("api_calls", [])) for p in gp)
            f.write(
                f"| [{group_name}](./{group_name}/README.md) "
                f"| `/{group_name}` "
                f"| {len(gp)} "
                f"| {api_cnt} |\n"
            )

        if undoc_pages:
            f.write(
                f"| [undocumented](./undocumented/README.md) "
                f"| — "
                f"| {len(undoc_pages)} "
                f"| — |\n"
            )

        f.write("\n---\n")
        f.write("_Generated by doc_writer_\n")


def _safe_endpoint_filename(endpoint: str) -> str:
    """Convert /api/v1/userinfo → api_v1_userinfo.md"""
    safe = endpoint.strip("/").replace("/", "_").replace("-", "_")
    safe = re.sub(r"[^\w]", "_", safe).strip("_") or "unknown"
    safe = re.sub(r"_+", "_", safe)
    return safe + ".md"


def _write_api_network_map(groups: Dict[str, List[dict]], output_root: str) -> None:
    """
    Write api_network_map.md — the master Frontend → Backend API dependency map.

    Shows every frontend page and ALL backend API endpoints it calls,
    equivalent to what you'd see in the browser Network tab when visiting each page.

    Format per page:
      ### `/route`
      Component: `Foo.vue`  |  URL: https://...
      ```
      GET     /api/userinfo       ← onMounted
      POST    /api/login          ← On form submit
      GET     /api/analytics      ← onMounted
      ```
    """
    out_path = os.path.join(output_root, "api_network_map.md")

    all_pages: List[dict] = []
    for group_name in sorted(groups.keys()):
        all_pages.extend(groups[group_name])

    total_pages = len(all_pages)
    total_calls = sum(len(p.get("api_calls", [])) for p in all_pages)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("# Frontend → Backend API Network Map\n\n")
        f.write(
            "> Shows which backend APIs each frontend page calls — equivalent to\n"
            "> what you see in the browser **Network** tab when visiting each page.\n\n"
        )
        f.write(f"**Pages**: {total_pages} | **Total API calls detected**: {total_calls}\n\n")
        f.write("---\n\n")

        # ── Quick-reference table ──────────────────────────────────────────────
        f.write("## Quick Reference\n\n")
        f.write("| Frontend Page | Frontend URL | APIs Called |\n")
        f.write("|--------------|-------------|-------------|\n")
        for page in all_pages:
            path        = page.get("path", "UNKNOWN")
            example_url = page.get("example_url") or "—"
            api_calls   = page.get("api_calls", [])
            inferred    = page.get("inferred", False)

            label = f"{path} _(inferred)_" if inferred else path

            unique_eps = list(dict.fromkeys(
                c.get("endpoint", "")
                for c in api_calls
                if c.get("endpoint") and c.get("endpoint") != "UNKNOWN"
            ))
            if unique_eps:
                ep_str = ", ".join(f"`{e}`" for e in unique_eps[:5])
                if len(unique_eps) > 5:
                    ep_str += f" _(+{len(unique_eps)-5} more)_"
            else:
                ep_str = "—"

            f.write(f"| `{label}` | `{example_url}` | {ep_str} |\n")

        f.write("\n---\n\n")

        # ── Detailed per-page network trace ────────────────────────────────────
        f.write("## Detailed API Calls per Page\n\n")
        for page in all_pages:
            path        = page.get("path", "UNKNOWN")
            component   = page.get("component", "UNKNOWN")
            example_url = page.get("example_url") or ""
            api_calls   = page.get("api_calls", [])
            inferred    = page.get("inferred", False)

            inferred_note = " _(route inferred from file path)_" if inferred else ""
            f.write(f"### `{path}`{inferred_note}\n\n")
            f.write(f"**Component**: `{component}`")
            if example_url:
                f.write(f" | **URL**: [{example_url}]({example_url})")
            f.write("\n\n")

            if api_calls:
                f.write("```\n")
                for call in api_calls:
                    method       = (call.get("method") or "?").upper().ljust(8)
                    endpoint     = call.get("endpoint", "UNKNOWN")
                    trigger_name = call.get("trigger_name", "")
                    trigger      = call.get("trigger", "")
                    purpose      = call.get("purpose", "")

                    if trigger_name:
                        hint = f"← {trigger_name}"
                    elif trigger == "lifecycle":
                        hint = "← on page load"
                    elif trigger == "event_handler":
                        hint = "← on user action"
                    elif trigger == "watcher":
                        hint = "← on data change"
                    elif purpose and purpose != "inline call":
                        hint = f"← {purpose}"
                    else:
                        hint = ""

                    f.write(f"{method} {endpoint}  {hint}\n".rstrip() + "\n")
                f.write("```\n\n")
            else:
                f.write("_No API calls detected._\n\n")

    print(f"  [MAP] API network map → {out_path}")




def _write_missing_apis(
    api_list: List[dict],
    output_root: str,
    config: "AIConfig",
    no_ai: bool = False,
) -> None:
    """
    Write undocumented API documentation.

    Creates:
      undocumented/missing_apis.md            — master index grouped by endpoint
      undocumented/apis/<endpoint>.md         — per-endpoint detail page (AI or skeleton)
      undocumented/unresolved_endpoints.md    — UNKNOWN/DYNAMIC endpoints with hints
    """
    undoc_dir = os.path.join(output_root, "undocumented")
    apis_dir  = os.path.join(undoc_dir, "apis")
    os.makedirs(apis_dir, exist_ok=True)

    # Split: real endpoints vs unresolved (UNKNOWN / DYNAMIC)
    real_apis      = [a for a in api_list if a.get("endpoint")
                      and a["endpoint"] != "UNKNOWN"
                      and not str(a.get("endpoint", "")).startswith("DYNAMIC: var ")]
    unresolved_raw = [a for a in api_list if not real_apis or a not in real_apis]

    # ── Write unresolved/unknown endpoints doc ────────────────────────────────
    if unresolved_raw:
        _write_unresolved_endpoints(unresolved_raw, undoc_dir, config, no_ai)

    if not real_apis:
        return

    # Group by endpoint → list of full usage dicts
    api_map: Dict[str, List[dict]] = defaultdict(list)
    for usage in real_apis:
        api_map[usage["endpoint"]].append(usage)

    sys_msg = pages_md_system()

    # ── Per-endpoint individual files ──────────────────────────────────────────
    endpoint_files: Dict[str, str] = {}   # endpoint -> filename (relative to apis/)

    for endpoint in sorted(api_map.keys()):
        usages    = api_map[endpoint]
        safe_name = _safe_endpoint_filename(endpoint)
        ep_path   = os.path.join(apis_dir, safe_name)
        endpoint_files[endpoint] = safe_name

        print(f"  [UNDOC API] {endpoint} -> undocumented/apis/{safe_name}")

        if no_ai or not config.use_ai:
            content = _skeleton_undoc_api(endpoint, usages)
        else:
            prompt = undocumented_api_prompt(endpoint, usages)
            try:
                content = call_ai(prompt, config, system=sys_msg, max_tokens=1400)
            except Exception as e:
                content = _skeleton_undoc_api(endpoint, usages)
            if not content or content.startswith("[AI failed"):
                content = _skeleton_undoc_api(endpoint, usages)
            if config.delay > 0:
                time.sleep(config.delay)

        with open(ep_path, "w", encoding="utf-8") as f:
            f.write(content)

    # ── Master missing_apis.md ─────────────────────────────────────────────────
    missing_path = os.path.join(undoc_dir, "missing_apis.md")
    with open(missing_path, "w", encoding="utf-8") as f:
        f.write("# API Endpoints Used by Frontend\n\n")
        f.write(
            "These API endpoints are called by frontend pages.\n\n"
        )
        f.write(f"**Total**: {len(api_map)} unique endpoints\n\n")
        if os.path.exists(os.path.join(undoc_dir, "unresolved_endpoints.md")):
            ur_count = len(set(a["endpoint"] for a in unresolved_raw))
            f.write(
                f"**Unresolved / Dynamic**: {ur_count} endpoint(s) — "
                f"see [unresolved_endpoints.md](unresolved_endpoints.md)\n\n"
            )
        f.write("| Endpoint | Methods | Pages Using It | Detail |\n")
        f.write("|----------|---------|----------------|--------|\n")
        for endpoint in sorted(api_map.keys()):
            usages       = api_map[endpoint]
            pages_using  = sorted(set(u["page_path"] for u in usages))
            methods      = sorted(set(u["method"] for u in usages))
            safe_name    = endpoint_files[endpoint]
            methods_str  = ", ".join(f"`{m}`" for m in methods)
            f.write(
                f"| `{endpoint}` | {methods_str} | {len(pages_using)} "
                f"| [detail](apis/{safe_name}) |\n"
            )
        f.write("\n---\n\n")

        for endpoint in sorted(api_map.keys()):
            usages      = api_map[endpoint]
            safe_name   = endpoint_files[endpoint]
            pages_using = sorted(set(u["page_path"] for u in usages))
            methods     = sorted(set(u["method"] for u in usages))

            f.write(f"## [`{endpoint}`](apis/{safe_name})\n\n")
            f.write(f"- **Methods**: {', '.join(f'`{m}`' for m in methods)}\n")
            f.write(f"- **Used by {len(pages_using)} page(s)**:\n\n")
            f.write("  | Page / Route | Method | Source | Transport |\n")
            f.write("  |-------------|--------|--------|-----------|\n")
            for u in usages:
                comp_note = (
                    f"via `{u['composable']}()`"
                    if u.get("composable")
                    else f"in `{u.get('called_from', '?')}`"
                )
                f.write(
                    f"  | `{u['page_path']}` | `{u['method']}` "
                    f"| {comp_note} | {u['via']} |\n"
                )
            f.write(f"\n[View full detail](apis/{safe_name})\n\n")

        f.write("---\n")

    # ── Append API cross-reference to undocumented README ─────────────────────
    readme_path = os.path.join(undoc_dir, "README.md")
    if os.path.exists(readme_path):
        with open(readme_path, "a", encoding="utf-8") as f:
            f.write("\n## API Endpoints Used by Frontend\n\n")
            f.write(
                f"{len(api_map)} API endpoint(s) are called by frontend pages. "
                f"See [missing_apis.md](missing_apis.md) for the full list "
                f"and individual files in [apis/](apis/) for per-endpoint detail.\n\n"
            )
            f.write("| Endpoint | Used by (pages) | Methods |\n")
            f.write("|----------|-----------------|---------|\n")
            for endpoint in sorted(api_map.keys()):
                usages      = api_map[endpoint]
                pages_using = sorted(set(u["page_path"] for u in usages))
                methods     = sorted(set(u["method"] for u in usages))
                safe_name   = endpoint_files[endpoint]
                methods_str = ", ".join(f"`{m}`" for m in methods)
                pages_str   = ", ".join(f"`{p}`" for p in pages_using[:3])
                if len(pages_using) > 3:
                    pages_str += f" (+{len(pages_using) - 3} more)"
                f.write(
                    f"| [`{endpoint}`](apis/{safe_name}) | {pages_str} | {methods_str} |\n"
                )
            f.write("\n")


def _write_unresolved_endpoints(
    unresolved: List[dict],
    undoc_dir: str,
    config: "AIConfig",
    no_ai: bool = False,
) -> None:
    """
    Write a deep-analysis report for UNKNOWN / DYNAMIC / unresolvable endpoints.

    For each unresolved call, shows:
      - Why it couldn't be resolved (variable URL, Vuex dispatch target missing, etc.)
      - What we do know (file it's in, transport, static prefix if DYNAMIC)
      - Env vars and URL constants available as context
      - Actionable steps to document it
      - AI-inferred endpoint pattern (when AI is enabled)
    """
    out_path = os.path.join(undoc_dir, "unresolved_endpoints.md")

    # Group by (endpoint, via, page) to avoid duplication
    _seen: set = set()
    unique: List[dict] = []
    for u in unresolved:
        key = (u.get("endpoint"), u.get("via"), u.get("page_path"),
               u.get("called_from"), u.get("vuex_action"))
        if key not in _seen:
            _seen.add(key)
            unique.append(u)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("# Unresolved / Dynamic API Endpoints\n\n")
        f.write(
            "> These API calls were detected in the frontend source code but the "
            "exact endpoint URL could **not** be determined statically.\n\n"
        )
        f.write(f"**Total unresolved calls**: {len(unique)}\n\n")
        f.write("---\n\n")

        # Collect all env/constants context from the entries
        all_env: Dict[str, str] = {}
        all_constants: Dict[str, str] = {}
        for u in unique:
            all_env.update(u.get("env_config") or {})
            all_constants.update(u.get("url_constants") or {})

        if all_env:
            f.write("## Known Base URLs (from `.env` files)\n\n")
            f.write("| Variable | Value |\n")
            f.write("|----------|-------|\n")
            for k, v in sorted(all_env.items()):
                f.write(f"| `{k}` | `{v}` |\n")
            f.write("\n")

        if all_constants:
            f.write("## Known URL Constants (from config/constants files)\n\n")
            f.write("| Constant | Value |\n")
            f.write("|----------|-------|\n")
            for k, v in sorted(all_constants.items()):
                f.write(f"| `{k}` | `{v}` |\n")
            f.write("\n")

        f.write("---\n\n## Unresolved Calls\n\n")

        for u in unique:
            endpoint    = u.get("endpoint", "UNKNOWN")
            via         = u.get("via", "direct")
            page_path   = u.get("page_path", "?")
            called_from = u.get("called_from", "?")
            method      = u.get("method", "?")
            vuex_action  = u.get("vuex_action")
            resolved_via = u.get("resolved_via")
            snippet      = u.get("context_snippet") or ""
            env_cfg      = u.get("env_config") or {}
            url_consts   = u.get("url_constants") or {}

            # Heading
            if endpoint.startswith("DYNAMIC:"):
                heading = f"DYNAMIC call in `{called_from}`"
            elif vuex_action:
                heading = f"Vuex dispatch: `{vuex_action}`"
            elif endpoint == "UNKNOWN":
                heading = f"UNKNOWN endpoint in `{called_from}` (page: `{page_path}`)"
            else:
                heading = f"`{endpoint}` (partial)"

            f.write(f"### {heading}\n\n")
            f.write("| Field | Value |\n")
            f.write("|-------|-------|\n")
            f.write(f"| **Raw endpoint** | `{endpoint}` |\n")
            f.write(f"| **Method** | `{method}` |\n")
            f.write(f"| **Transport** | `{via}` |\n")
            f.write(f"| **Page route** | `{page_path}` |\n")
            f.write(f"| **Source file** | `{called_from}` |\n")
            if vuex_action:
                f.write(f"| **Vuex action key** | `{vuex_action}` |\n")
            if resolved_via:
                f.write(f"| **Resolved from** | `{resolved_via}` |\n")

            # ── AI deep-resolution ────────────────────────────────────────────
            # Only attempt when: AI is on + there's a code snippet + it's DYNAMIC/UNKNOWN
            should_ai_infer = (
                not no_ai
                and config.use_ai
                and snippet
                and (endpoint.startswith("DYNAMIC:") or endpoint == "UNKNOWN")
                and via not in ("vuex_dispatch",)
            )
            if should_ai_infer:
                try:
                    infer_prompt = resolve_dynamic_endpoint_prompt(
                        raw_endpoint    = endpoint,
                        method          = method,
                        called_from     = called_from,
                        context_snippet = snippet,
                        env_config      = env_cfg,
                        url_constants   = url_consts,
                    )
                    ai_response = call_ai(infer_prompt, config, max_tokens=400)
                    if ai_response and not ai_response.startswith("[AI failed"):
                        f.write("\n**AI-Inferred Endpoint:**\n\n")
                        f.write(ai_response.strip())
                        f.write("\n\n")
                        if config.delay > 0:
                            time.sleep(config.delay)
                except Exception:
                    pass  # AI inference is best-effort only

            # ── Show code snippet ─────────────────────────────────────────────
            if snippet:
                f.write("\n**Code context:**\n\n")
                f.write(f"```js\n{snippet}\n```\n\n")

            # ── Resolution hints ──────────────────────────────────────────────
            f.write("**How to resolve:**\n\n")
            if vuex_action:
                parts = vuex_action.split("/")
                mod   = parts[0] if len(parts) > 1 else "root"
                act   = parts[-1]
                f.write(
                    f"1. Find the Vuex store module `{mod}` "
                    f"(e.g. `store/modules/{mod}.js`)\n"
                    f"2. Locate the action `{act}` and its API calls\n"
                    f"3. Document that endpoint in the backend docs\n\n"
                )
            elif endpoint.startswith("DYNAMIC:"):
                raw = endpoint[len("DYNAMIC:"):].strip()
                f.write(
                    f"1. Open `{called_from}` and locate the call using: `{raw}`\n"
                    f"2. Check the variable(s) or expression used to build the URL\n"
                    f"3. Common patterns: `const url = '/api/...'` or "
                    f"`\\`/api/${{id}}\\``\n"
                )
                if all_env:
                    env_list = ", ".join(f"`{k}`" for k in all_env)
                    f.write(
                        f"4. The base URL may come from env var(s): {env_list}\n"
                    )
                f.write("\n")
            elif via == "graphql":
                f.write(
                    "1. This is a GraphQL call — check your Apollo/GraphQL schema\n"
                    "2. Locate the query/mutation definition file\n"
                    "3. Document the operation in the GraphQL API docs\n\n"
                )
            elif via == "websocket":
                f.write(
                    "1. This is a WebSocket connection — check WebSocket server config\n"
                    "2. Locate the handler on the backend\n"
                    "3. Document the WS events/messages in the API docs\n\n"
                )
            else:
                f.write(
                    f"1. Open `{called_from}` and search for the HTTP call\n"
                    "2. Trace the URL variable backward to its declaration\n\n"
                )

            f.write("---\n\n")


def _skeleton_undoc_api(endpoint: str, usages: List[dict]) -> str:
    """No-AI fallback: generate structured markdown for an undocumented API endpoint."""
    pages_using = sorted(set(u["page_path"] for u in usages))
    methods     = sorted(set(u["method"] for u in usages))
    methods_str = ", ".join(f"`{m}`" for m in methods)

    # Collect env / URL constants context across all usages
    all_env: Dict[str, str] = {}
    all_constants: Dict[str, str] = {}
    for u in usages:
        all_env.update(u.get("env_config") or {})
        all_constants.update(u.get("url_constants") or {})

    # Detect if this is a GraphQL or WebSocket endpoint
    is_graphql  = endpoint.startswith("graphql:")
    is_ws       = endpoint.startswith("ws://") or endpoint.startswith("wss://")
    is_dynamic  = endpoint.startswith("DYNAMIC:")

    lines = [
        f"# API Endpoint: `{endpoint}`\n\n",
        f"> This endpoint is called by the frontend.\n\n",
        "## Summary\n\n",
        "| Field | Value |\n",
        "|-------|-------|\n",
        f"| **Endpoint** | `{endpoint}` |\n",
        f"| **HTTP Methods** | {methods_str} |\n",
        f"| **Used by** | {len(pages_using)} page(s) |\n",
    ]

    if is_graphql:
        op_name = endpoint[len("graphql:"):]
        lines.append(f"| **Type** | GraphQL |\n")
        lines.append(f"| **Operation** | `{op_name}` |\n")
    elif is_ws:
        lines.append("| **Type** | WebSocket |\n")
    elif is_dynamic:
        lines.append("| **Type** | Dynamic (runtime-computed URL) |\n")

    lines.append("\n")

    # ── Base URL context ───────────────────────────────────────────────────────
    if all_env or all_constants:
        lines.append("## Base URL Context\n\n")
        if all_env:
            lines.append("**From `.env` files:**\n\n")
            lines.append("| Variable | Value |\n")
            lines.append("|----------|-------|\n")
            for k, v in sorted(all_env.items()):
                lines.append(f"| `{k}` | `{v}` |\n")
            lines.append("\n")
        if all_constants:
            lines.append("**From config/constants files:**\n\n")
            lines.append("| Constant | Value |\n")
            lines.append("|----------|-------|\n")
            for k, v in sorted(all_constants.items()):
                lines.append(f"| `{k}` | `{v}` |\n")
            lines.append("\n")

    # ── Where it is used ──────────────────────────────────────────────────────
    lines.append("## Where It Is Used\n\n")
    lines.append("| Page / Route | Method | Source | Transport |\n")
    lines.append("|-------------|--------|--------|-----------|\n")
    for u in usages:
        comp_note = (
            f"via `{u['composable']}()`"
            if u.get("composable")
            else f"in `{u.get('called_from', 'UNKNOWN')}`"
        )
        resolved = f" ← `{u['resolved_via']}`" if u.get("resolved_via") else ""
        lines.append(
            f"| `{u['page_path']}` | `{u['method']}` "
            f"| {comp_note}{resolved} | {u['via']} |\n"
        )

    lines.append("\n## Pages Detail\n\n")
    for page_path in pages_using:
        page_usages = [u for u in usages if u["page_path"] == page_path]
        lines.append(f"### `{page_path}`\n\n")
        for u in page_usages:
            comp_note = (
                f"Called via composable `{u['composable']}()`"
                if u.get("composable")
                else f"Called directly in `{u.get('called_from', 'UNKNOWN')}`"
            )
            vuex_note = (
                f" (Vuex action: `{u['vuex_action']}`)"
                if u.get("vuex_action") else ""
            )
            lines.append(
                f"- {comp_note}{vuex_note} | Method: `{u['method']}` | Transport: `{u['via']}`\n"
            )
        lines.append("\n")

    # ── Infer purpose from endpoint name ──────────────────────────────────────
    ep_lower = endpoint.lower()
    inferences = []
    if any(x in ep_lower for x in ["list", "all", "index"]):
        inferences.append("Returns a list/collection of resources")
    if any(x in ep_lower for x in ["detail", "show", "get", "info", "view"]):
        inferences.append("Returns detailed information about a specific resource")
    if any(x in ep_lower for x in ["create", "add", "store", "new"]):
        inferences.append("Creates a new resource")
    if any(x in ep_lower for x in ["update", "edit", "modify", "save"]):
        inferences.append("Updates an existing resource")
    if any(x in ep_lower for x in ["delete", "remove", "destroy"]):
        inferences.append("Deletes a resource")
    if any(x in ep_lower for x in ["auth", "login", "logout"]):
        inferences.append("Handles authentication")
    if any(x in ep_lower for x in ["user", "profile", "account"]):
        inferences.append("Handles user/account data")
    if any(x in ep_lower for x in ["report", "stats", "analytics", "summary", "dashboard"]):
        inferences.append("Provides reporting or analytics data")
    if any(x in ep_lower for x in ["upload", "import", "export", "download"]):
        inferences.append("Handles file upload/download or data import/export")
    if is_graphql:
        inferences.append("GraphQL operation — check the server-side resolver")
    if is_ws:
        inferences.append("WebSocket connection — check the WS server handler")

    lines.append("## How It Can Be Used\n\n")
    lines.append(f"_Based on endpoint `{endpoint}`, this API likely:_\n\n")
    if inferences:
        for inf in inferences:
            lines.append(f"- {inf}\n")
    else:
        lines.append(f"- Handles `{methods[0] if methods else 'HTTP'}` requests\n")
        lines.append(f"- Used by {len(pages_using)} frontend page(s)\n")

    if not is_graphql and not is_ws:
        lines.append(
            f"\n**Example call (axios)**:\n"
            f"```js\n"
            f"// {methods[0] if methods else 'GET'} {endpoint}\n"
            f"const response = await axios.{(methods[0] if methods else 'GET').lower()}"
            f"('{endpoint}');\n"
            f"```\n\n"
        )

    lines.append(
        "## Notes\n\n"
        "> To get more details about this endpoint:\n"
        "> 1. Locate the backend controller/route that handles this path\n"
        "> 2. Check the backend API documentation for this route\n\n"
        "---\n"
    )
    return "".join(lines)


def _skeleton_page(page: dict) -> str:
    """No-AI fallback: produce network-trace-focused markdown from extracted page data."""
    from frontend.detect_pages import \
        _build_example_url  # lazy import to avoid circular

    path      = page.get("path", "UNKNOWN")
    component = page.get("component", "UNKNOWN")
    comp_file = page.get("component_file")
    example_url = page.get("example_url")
    api_calls = page.get("api_calls", [])
    unknowns  = page.get("unknowns", [])
    inferred  = page.get("inferred", False)

    if not example_url or str(example_url).strip() in ("N/A", "None", "null", ""):
        example_url = _build_example_url(path)

    comp_file_display = f"`{comp_file}`" if comp_file else "_not found on disk_"
    inferred_row = "| **Route Source** | _Inferred from file path_ |\n" if inferred else ""

    if example_url:
        url_row     = f"| **Frontend URL** | `{example_url}` |\n"
        url_callout = f"> To verify: **[{example_url}]({example_url})**\n\n"
    else:
        url_row     = "| **Frontend URL** | _Route not mapped_ |\n"
        url_callout = "> Route has no URL mapping — component may be a modal or child view.\n\n"

    # ── API network-trace table ───────────────────────────────────────────────
    api_rows = []
    for i, call in enumerate(api_calls, 1):
        method       = (call.get("method") or "?").upper()
        endpoint     = call.get("endpoint", "UNKNOWN")
        trigger_name = call.get("trigger_name", "")
        trigger      = call.get("trigger", "unknown")
        composable   = call.get("composable")
        via_child    = call.get("via_child")
        purpose      = call.get("purpose", "")
        comment      = call.get("comment", "")
        dynamic      = call.get("dynamic", False)

        if trigger_name:
            when = f"`{trigger_name}`"
        elif trigger == "lifecycle":
            when = "On page load"
        elif trigger == "event_handler":
            when = "On user action"
        elif trigger == "watcher":
            when = "On data change"
        elif trigger == "function_call":
            fn = call.get("function_name", "")
            when = f"In `{fn}()`" if fn else "On function call"
        else:
            when = "On function call"

        if composable:
            when += f" via `{composable}()`"
        if via_child:
            when += f" (child: `{via_child}`)"
        if dynamic:
            endpoint += " ⚡"

        short_purpose = purpose if purpose and purpose != "inline call" else (comment or "—")
        api_rows.append(
            f"| {i} | `{method}` | `{endpoint}` | {when} | {short_purpose} |"
        )

    if api_rows:
        api_md = (
            "| # | Method | Endpoint | When Called | Purpose |\n"
            "|---|--------|----------|-------------|--------|\n"
            + "\n".join(api_rows)
        )
    elif comp_file:
        api_md = "_No API calls detected in this component._"
    else:
        api_md = "_Component source not found — could not scan for API calls._"

    unknowns_md = ""
    if unknowns:
        unknowns_md = "\n## Notes\n\n" + "\n".join(f"- {u}" for u in unknowns) + "\n\n"

    return (
        f"# `{path}`\n\n"
        f"| Field | Value |\n"
        f"|-------|-------|\n"
        f"| **Component** | `{component}` |\n"
        f"| **Source File** | {comp_file_display} |\n"
        f"{inferred_row}"
        f"{url_row}\n"
        f"{url_callout}"
        f"## What This Page Does\n\n"
        f"_Enable AI to generate a description of this page's business purpose._\n\n"
        f"## APIs Called by This Page\n\n"
        f"{api_md}\n\n"
        f"{unknowns_md}"
        f"---"
    )
