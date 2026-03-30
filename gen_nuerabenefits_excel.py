"""
Generates nuerabenefits_backend.xlsx from the markdown docs.
One sheet, same style as Book.xlsx reference (light-blue / white alternating rows).
Includes both backend domain sheets and frontend route-group sheets.
"""

import re
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE_DIR      = Path(r'D:\CloudTech_main\Doc_writer\doc_output\nuerabenefits\docs\backend')
FRONTEND_DIR  = Path(r'D:\CloudTech_main\Doc_writer\doc_output\nuerabenefits\docs\frontend')
OUT_FILE      = Path(r'D:\CloudTech_main\Doc_writer\nuerabenefits_api_detail.xlsx')
TODAY         = date.today().strftime('%Y-%m-%d')

# ── palette ──────────────────────────────────────────────────────────────────
C = {
    'title_bg'   : 'FF1B2A4A',
    'meta_label' : 'FF2563EB',
    'meta_value' : 'FF2D3A58',
    'hdr_bg'     : 'FF243B6E',
    'sec_bg'     : 'FF1B2A4A',
    'row_even'   : 'FFEFF6FF',  # light blue
    'row_odd'    : 'FFFFFFFF',  # white
    'db_even'    : 'FFE8F0FE',
    'db_odd'     : 'FFF1F5F9',
    'resp_bg'    : 'FFFAFAFA',
    'stat_bg'    : 'FFF0F4FF',
    'get_bg'     : 'FFDCFCE7',
    'post_bg'    : 'FFEDE9FE',
    'put_bg'     : 'FFFEF3C7',
    'del_bg'     : 'FFFFE4E6',
    'white'      : 'FFFFFFFF',
    'dark'       : 'FF1E293B',
    # Frontend additions (teal palette)
    'fe_title'   : 'FF134E4A',
    'fe_hdr'     : 'FF0F766E',
    'fe_meta'    : 'FF0D9488',
    'fe_row_even': 'FFF0FDFA',
    'fe_row_odd' : 'FFFFFFFF',
    'text_blue'  : 'FF1D4ED8',
}

def fill(hex_c):
    return PatternFill(start_color=hex_c, end_color=hex_c, fill_type='solid')

def font(bold=False, color='FF000000', size=9, italic=False):
    return Font(bold=bold, color=color, size=size, name='Calibri', italic=italic)

def align(h='left', v='top', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

_thin = Side(style='thin', color='FFCBD5E1')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def method_fill(m):
    m = m.upper()
    return fill({'GET': C['get_bg'], 'POST': C['post_bg'],
                 'PUT': C['put_bg'],  'PATCH': C['put_bg'],
                 'DELETE': C['del_bg']}.get(m, C['row_even']))

def default_statuses(m):
    m = m.upper()
    if m == 'GET':    return '200 OK\n422 Validation Error\n404 Not Found'
    if m == 'POST':   return '201 Created\n200 OK\n422 Validation Error\n404 Not Found'
    if m in ('PUT','PATCH'): return '200 OK\n422 Validation Error\n404 Not Found'
    if m == 'DELETE': return '200 OK\n404 Not Found'
    return '200 OK'

# ── markdown parsers ──────────────────────────────────────────────────────────

def parse_api_md(path: Path):
    """Returns list of {name, method, url, controller}."""
    if not path.exists():
        return []
    text = path.read_text(encoding='utf-8', errors='ignore')
    results = []
    for block in re.split(r'\n## ', text)[1:]:
        lines  = block.strip().splitlines()
        name   = lines[0].strip()
        ep_raw = ''
        ctrl   = ''
        for line in lines:
            m = re.search(r'\*\*Endpoint\*\*\s*:\s*`([^`]+)`', line)
            if m: ep_raw = m.group(1).strip()
            m = re.search(r'\*\*Controller\*\*\s*:\s*`([^`]+)`', line)
            if m: ctrl = m.group(1).strip()
        if not name or not ep_raw:
            continue
        parts  = ep_raw.split(' ', 1)
        method = parts[0].upper() if len(parts) == 2 else 'GET'
        url    = parts[1] if len(parts) == 2 else ep_raw
        results.append({'name': name, 'method': method, 'url': url,
                        'endpoint': ep_raw, 'controller': ctrl})
    return results


def _get_section(text: str, name: str) -> str:
    """Return the content block of the ## name section."""
    for block in re.split(r'\n## ', text)[1:]:
        if block.strip().splitlines()[0].strip().lower() == name.lower():
            return block
    return ''


def parse_business_md(path: Path, ep_name: str):
    if not path.exists():
        return {}
    text    = path.read_text(encoding='utf-8', errors='ignore')
    section = _get_section(text, ep_name)
    if not section:
        return {}

    def sub(heading):
        m = re.search(rf'### {heading}\s*\n+(.*?)(?=\n###|\Z)', section, re.DOTALL)
        return m.group(1).strip() if m else ''

    auth = re.search(r'\*\*Auth Required\*\*\s*\|\s*([^\n|]+)', section)
    http = re.search(r'\*\*HTTP Method\*\*\s*\|\s*([^\n|]+)', section)

    return {
        'auth'    : auth.group(1).strip() if auth else 'Yes (verify.internal.token)',
        'method'  : http.group(1).strip().upper() if http else '',
        'purpose' : sub('Purpose'),
        'biz'     : sub('Business Logic'),
        'db'      : sub('Database Operations'),
    }


def parse_responses_md(path: Path, method: str, url: str):
    if not path.exists():
        return {}
    text = path.read_text(encoding='utf-8', errors='ignore')
    # Try matching "## METHOD /path"
    url_stem = url.split('{')[0].rstrip('/')
    target = ''
    for block in re.split(r'\n## ', text)[1:]:
        header = block.strip().splitlines()[0].strip()
        if method.upper() in header.upper() and url_stem in header:
            target = block
            break
    if not target:
        for block in re.split(r'\n## ', text)[1:]:
            header = block.strip().splitlines()[0].strip()
            if url_stem and url_stem in header:
                target = block
                break
    if not target:
        return {}
    m = re.search(r'```json\s*(.*?)\s*```', target, re.DOTALL)
    return {'response': m.group(1).strip()} if m else {}


# ── Frontend page parser ──────────────────────────────────────────────────────

def parse_frontend_md(path: Path) -> list:
    """Parse a frontend page .md file. Returns list of page dicts."""
    if not path.exists():
        return []
    text = path.read_text(encoding='utf-8', errors='ignore')
    pages = []
    for sec in re.split(r'\n(?=# (?:Page|`/))', text):
        sec = sec.strip()
        if not sec:
            continue
        route = ''
        hm = re.match(r'^#\s+(?:Page:\s*)?`([^`]+)`', sec)
        if hm:
            route = hm.group(1).strip()

        def _field(label):
            m = re.search(
                rf'\|\s*\*\*{re.escape(label)}\*\*\s*\|\s*`?([^|\n`]+)`?\s*\|', sec)
            return m.group(1).strip() if m else ''

        component   = _field('Component')
        source_file = _field('Source file')
        layout      = _field('Layout')
        example_url = _field('Example URL')

        # Children
        children = []
        cm = re.search(r'## Child Components\n+(.*?)(?=\n##|\Z)', sec, re.DOTALL)
        if cm:
            block = cm.group(1).strip()
            if block.lower() not in ('none', 'none detected',
                                     'none — no imported or template sub-components detected'):
                children = [l.lstrip('- ').strip().strip('`')
                            for l in block.splitlines() if l.strip().startswith('-')]

        # Composables
        composables = []
        comp_m = re.search(r'## Composables Used\n+(.*?)(?=\n##|\Z)', sec, re.DOTALL)
        if comp_m:
            block = comp_m.group(1).strip()
            if block.lower() not in ('none', 'none detected'):
                composables = [l.lstrip('- ').strip().strip('`').rstrip('()')
                               for l in block.splitlines() if l.strip().startswith('-')]

        # API calls
        api_calls = []
        api_m = re.search(r'## Backend API Dependencies\n+(.*?)(?=\n##|\Z)', sec, re.DOTALL)
        if api_m:
            block = api_m.group(1).strip()
            if block.lower() not in ('none', 'none detected'):
                for row in re.finditer(
                    r'\|\s*`([A-Z]+)`\s*\|\s*`([^`]+)`\s*\|([^|]+)\|([^|]+)\|', block):
                    api_calls.append({'method': row.group(1).strip(),
                                      'endpoint': row.group(2).strip(),
                                      'source': row.group(3).strip(),
                                      'via': row.group(4).strip()})
                if not api_calls:
                    for ep_line in re.finditer(
                            r'-\s*Endpoint:\s*`([^`]+)`.*?Method:\s*([A-Z]+)',
                            block, re.DOTALL):
                        api_calls.append({'method': ep_line.group(2).strip(),
                                          'endpoint': ep_line.group(1).strip(),
                                          'source': '', 'via': 'axios'})

        state_m = re.search(r'## State Management\n+(.*?)(?=\n##|\Z)', sec, re.DOTALL)
        state_mgmt = state_m.group(1).strip() if state_m else ''
        if state_mgmt.lower() in ('none', 'none detected', '—'):
            state_mgmt = '—'

        warn_m = re.search(r'## Warnings\n+(.*?)(?=\n##|\Z)', sec, re.DOTALL)
        warnings = warn_m.group(1).strip() if warn_m else ''
        if warnings.lower() in ('none', '_none_'):
            warnings = ''

        pages.append({'route': route or 'UNKNOWN', 'component': component,
                      'source_file': source_file, 'layout': layout,
                      'example_url': example_url, 'children': children,
                      'composables': composables, 'api_calls': api_calls,
                      'state_mgmt': state_mgmt, 'warnings': warnings})
    return pages


def load_frontend_group(group_dir: Path) -> list:
    """Load all page .md files from a frontend group folder (skip README.md)."""
    pages = []
    for f in sorted(group_dir.iterdir()):
        if f.name.lower() == 'readme.md' or f.suffix != '.md':
            continue
        pages.extend(parse_frontend_md(f))
    return pages


# ── Frontend Excel sheet builder ──────────────────────────────────────────────

FE_COLS = [
    ('#',                                   4),
    ('Route Path',                         30),
    ('Component',                          20),
    ('Layout',                             14),
    ('Example URL',                        32),
    ('Child Components',                   28),
    ('Composables',                        22),
    ('API Dependencies\n(Endpoint | Method)', 40),
    ('State Management',                   22),
    ('Source File',                        36),
    ('Warnings',                           28),
]
FE_N = len(FE_COLS)


def _build_fe_sheet(wb: openpyxl.Workbook, group_name: str,
                    pages: list) -> None:
    safe = re.sub(r'[\\/*?:\[\]]', '', f'FE_{group_name}')[:31]
    ws = wb.create_sheet(title=safe)
    ws.freeze_panes = 'A4'

    from openpyxl.utils import get_column_letter
    for idx, (_, w) in enumerate(FE_COLS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=FE_N)
    t = ws.cell(row=1, column=1, value=f'🖥️  Frontend Pages — /{group_name}')
    t.fill = fill(C['fe_title']); t.font = Font(bold=True, color=C['white'], size=13, name='Calibri')
    t.alignment = align('left', 'center', wrap=False)
    ws.row_dimensions[1].height = 28

    # Meta
    meta = [('Group:', f'/{group_name}'), ('Pages:', str(len(pages))),
            ('Last Updated:', TODAY)]
    col = 1
    for lbl, val in meta:
        lc = ws.cell(row=2, column=col, value=lbl)
        lc.fill = fill(C['fe_meta']); lc.font = Font(bold=True, color=C['white'], size=10, name='Calibri')
        lc.alignment = align('right', 'center', wrap=False)
        vc = ws.cell(row=2, column=col + 1, value=val)
        vc.fill = fill(C['meta_value']); vc.font = Font(color=C['white'], size=10, name='Calibri')
        vc.alignment = align('left', 'center', wrap=False)
        col += 2
    for c_idx in range(col, FE_N + 1):
        ws.cell(row=2, column=c_idx).fill = fill(C['meta_value'])
    ws.row_dimensions[2].height = 18

    # Headers
    for idx, (hdr, _) in enumerate(FE_COLS, 1):
        c = ws.cell(row=3, column=idx, value=hdr)
        c.fill = fill(C['fe_hdr']); c.font = Font(bold=True, color=C['white'], size=10, name='Calibri')
        c.alignment = align('center', 'center'); c.border = BORDER
    ws.row_dimensions[3].height = 36

    if not pages:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=FE_N)
        c = ws.cell(row=4, column=1, value='No page data found for this group.')
        c.fill = fill(C['fe_row_even']); c.font = Font(italic=True, color='FF94A3B8', size=9, name='Calibri')
        c.alignment = align('center'); ws.row_dimensions[4].height = 24
        return

    for i, page in enumerate(pages):
        row    = i + 4
        is_e   = (i % 2 == 0)
        bg     = C['fe_row_even'] if is_e else C['fe_row_odd']
        db_bg  = C['db_even'] if is_e else C['db_odd']

        api_text = '\n'.join(
            f"{c.get('method','?')}  {c.get('endpoint','?')}"
            for c in page.get('api_calls', [])
        ) or '—'
        children_text   = '\n'.join(page.get('children', [])) or '—'
        composable_text = '\n'.join(page.get('composables', [])) or '—'

        ws.row_dimensions[row].height = max(60, 15 * max(
            api_text.count('\n') + 1,
            children_text.count('\n') + 1,
            composable_text.count('\n') + 1, 2))

        def _c(col, val, cbg=None, bold=False, h='left', color=None):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill(cbg or bg)
            cell.font = Font(bold=bold, color=color or C['dark'], size=9, name='Calibri')
            cell.alignment = Alignment(horizontal=h, vertical='top', wrap_text=True)
            cell.border = BORDER

        _c(1,  i + 1,                           cbg=C['fe_title'], bold=True, h='center', color=C['white'])
        _c(2,  page.get('route', 'UNKNOWN'),     bold=True, color=C['text_blue'])
        _c(3,  page.get('component', '—'))
        _c(4,  page.get('layout', '—') or '—')
        _c(5,  page.get('example_url', '—') or '—')
        _c(6,  children_text,                    cbg=C['resp_bg'])
        _c(7,  composable_text,                  cbg=C['resp_bg'])
        _c(8,  api_text,                         cbg=db_bg)
        _c(9,  page.get('state_mgmt', '—') or '—')
        _c(10, page.get('source_file', '—') or '—', color='FF64748B')
        _c(11, page.get('warnings', '') or '—',  cbg=C['stat_bg'])


def _build_fe_index_sheet(wb: openpyxl.Workbook, group_summary: list) -> None:
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet(title='🖥️ FE Index')
    ws.freeze_panes = 'A4'

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    t = ws.cell(row=1, column=1, value='🖥️  Frontend — Page Groups Index')
    t.fill = fill(C['fe_title']); t.font = Font(bold=True, color=C['white'], size=14, name='Calibri')
    t.alignment = align('left', 'center', wrap=False)
    ws.row_dimensions[1].height = 32

    for col, (lbl, val) in enumerate([('Project:', 'nuerabenefits'),
                                       ('Generated:', TODAY),
                                       ('Groups:', str(len(group_summary)))], start=1):
        lc = ws.cell(row=2, column=col * 2 - 1, value=lbl)
        lc.fill = fill(C['fe_meta']); lc.font = Font(bold=True, color=C['white'], size=10, name='Calibri')
        lc.alignment = align('right', 'center', wrap=False)
        vc = ws.cell(row=2, column=col * 2, value=val)
        vc.fill = fill(C['meta_value']); vc.font = Font(color=C['white'], size=10, name='Calibri')
        vc.alignment = align('left', 'center', wrap=False)
    ws.row_dimensions[2].height = 18

    headers = ['#', 'Group / Prefix', 'Pages', 'API Calls', 'Sheet']
    widths  = [4,    24,               10,       12,           20]
    for idx, (hdr, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=idx, value=hdr)
        c.fill = fill(C['fe_hdr']); c.font = Font(bold=True, color=C['white'], size=10, name='Calibri')
        c.alignment = align('center', 'center'); c.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.row_dimensions[3].height = 28

    for i, (group, page_count, api_count) in enumerate(group_summary, 1):
        row = i + 3
        bg  = C['fe_row_even'] if i % 2 == 0 else C['fe_row_odd']
        ws.row_dimensions[row].height = 18
        sheet_name = re.sub(r'[\\/*?:\[\]]', '', f'FE_{group}')[:31]
        for idx, val in enumerate([i, f'/{group}', page_count, api_count, sheet_name], 1):
            c = ws.cell(row=row, column=idx, value=val)
            c.fill = fill(bg)
            c.font = Font(bold=(idx == 2), size=9, name='Calibri',
                          color=C['text_blue'] if idx == 2 else C['dark'])
            c.alignment = align('center' if idx in (1, 3, 4) else 'left')
            c.border = BORDER


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel():
    modules = sorted(
        [d for d in BASE_DIR.iterdir() if d.is_dir() and (d / 'api.md').exists()],
        key=lambda p: p.name
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'NueraBenefits Backend'

    # Column widths
    widths = [5, 28, 42, 10, 44, 42, 52, 28, 28, 16, 15, 44, 24, 26, 12, 46, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ── Row 1: title ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:Q1')
    t = ws['A1']
    t.value     = '🔌  API Detail Sheet — NUERABENEFITS BACKEND'
    t.fill      = fill(C['title_bg'])
    t.font      = Font(bold=True, color=C['white'], size=14, name='Calibri')
    t.alignment = align('center', 'center')
    ws.row_dimensions[1].height = 32

    # ── Row 2: metadata ───────────────────────────────────────────────────────
    meta = [
        ('A2', 'Module:',       C['meta_label'], True),
        ('B2', 'nuerabenefits', C['meta_value'], False),
        ('C2', 'Last Updated:', C['meta_label'], True),
        ('D2', TODAY,           C['meta_value'], False),
        ('E2', 'Status:',       C['meta_label'], True),
        ('F2', 'Active',        C['meta_value'], False),
        ('G2', 'API Version:',  C['meta_label'], True),
        ('H2', 'V1 / V2 / V3', C['meta_value'], False),
    ]
    for coord, val, clr, bold in meta:
        c           = ws[coord]
        c.value     = val
        c.fill      = fill(clr)
        c.font      = Font(bold=bold, color=C['white'], size=10, name='Calibri')
        c.alignment = align('left', 'center', wrap=False)
    ws.row_dimensions[2].height = 20

    # ── Row 3: column headers ─────────────────────────────────────────────────
    hdrs = ['#', 'Function / Endpoint Name', 'Purpose / Trigger',
            'HTTP\nMethod', 'Endpoint URL', 'DB Table /\nOperation',
            'Business Rules /\nNotes', 'Open Questions', 'Answer / Decision',
            'Answered By', 'Date Answered', 'Response Fields',
            'Status Codes', 'Auth / Middleware', 'Priority',
            'Owner\n(Controller)', 'Last Updated']
    for col, label in enumerate(hdrs, 1):
        c           = ws.cell(row=3, column=col, value=label)
        c.fill      = fill(C['hdr_bg'])
        c.font      = Font(bold=True, color=C['white'], size=10, name='Calibri')
        c.alignment = align('center', 'center')
        c.border    = BORDER
    ws.row_dimensions[3].height = 36

    # ── Data rows ─────────────────────────────────────────────────────────────
    row_num  = 4
    glob_idx = 1

    for mod_dir in modules:
        mod  = mod_dir.name
        eps  = parse_api_md(mod_dir / 'api.md')
        if not eps:
            continue

        # section header
        ws.merge_cells(f'A{row_num}:Q{row_num}')
        s           = ws[f'A{row_num}']
        s.value     = f'  ◆  {mod.upper()}'
        s.fill      = fill(C['sec_bg'])
        s.font      = Font(bold=True, color=C['white'], size=10, name='Calibri')
        s.alignment = align('left', 'center', wrap=False)
        ws.row_dimensions[row_num].height = 22
        row_num += 1

        for i, ep in enumerate(eps):
            biz  = parse_business_md(mod_dir / 'business.md', ep['name'])
            resp = parse_responses_md(mod_dir / 'responses.md', ep['method'], ep['url'])

            method   = biz.get('method') or ep['method']
            purpose  = biz.get('purpose') or f'Handles {ep["name"]} operation.'
            biz_text = biz.get('biz', '')
            db_text  = biz.get('db', '')
            auth     = biz.get('auth', 'Yes (verify.internal.token)')
            resp_txt = resp.get('response', '{\n  "data": "UNKNOWN"\n}')
            statuses = default_statuses(method)

            is_even  = (i % 2 == 0)
            rc       = C['row_even'] if is_even else C['row_odd']
            dc       = C['db_even']  if is_even else C['db_odd']

            def wc(col, val, bg, bold=False, h='left', v='top', wrap=True):
                c           = ws.cell(row=row_num, column=col, value=val)
                c.fill      = fill(bg)
                c.font      = Font(bold=bold, color=C['dark'], size=9, name='Calibri')
                c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
                c.border    = BORDER

            # A – index
            idx_c           = ws.cell(row=row_num, column=1, value=glob_idx)
            idx_c.fill      = fill(C['sec_bg'])
            idx_c.font      = Font(bold=True, color=C['white'], size=9, name='Calibri')
            idx_c.alignment = align('center', 'center', wrap=False)
            idx_c.border    = BORDER

            # B – name
            wc(2,  ep['name'],     rc,            bold=True)
            # C – purpose
            wc(3,  purpose,        rc)
            # D – method
            mc           = ws.cell(row=row_num, column=4, value=method)
            mc.fill      = method_fill(method)
            mc.font      = Font(bold=True, color=C['dark'], size=9, name='Calibri')
            mc.alignment = align('center', 'center', wrap=False)
            mc.border    = BORDER
            # E – endpoint URL
            wc(5,  ep['endpoint'], rc,            bold=True)
            # F – DB ops
            wc(6,  db_text,        dc)
            # G – business rules
            wc(7,  biz_text,       dc)
            # H,I,J,K – open questions (blank)
            for col in (8, 9, 10, 11):
                wc(col, '', rc)
            # L – response fields
            wc(12, resp_txt,       C['resp_bg'])
            # M – status codes
            wc(13, statuses,       C['stat_bg'])
            # N – auth
            wc(14, auth,           rc)
            # O – priority
            wc(15, '—',            rc, h='center')
            # P – controller
            wc(16, ep['controller'], rc)
            # Q – last updated
            wc(17, TODAY,          rc, h='center')

            ws.row_dimensions[row_num].height = 65
            row_num  += 1
            glob_idx += 1

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:Q{row_num - 1}'

    # ── Frontend sheets ───────────────────────────────────────────────────────
    fe_summary = []
    if FRONTEND_DIR.is_dir():
        fe_groups = sorted([
            d for d in FRONTEND_DIR.iterdir()
            if d.is_dir() and d.name != 'undocumented'
        ], key=lambda p: p.name)
        print(f'  Frontend groups found: {len(fe_groups)}')
        for gdir in fe_groups:
            pages      = load_frontend_group(gdir)
            total_api  = sum(len(p.get('api_calls', [])) for p in pages)
            fe_summary.append((gdir.name, len(pages), total_api))
            _build_fe_sheet(wb, gdir.name, pages)
            print(f'    FE Group: {gdir.name}  ({len(pages)} pages, {total_api} API calls)')
        if fe_summary:
            _build_fe_index_sheet(wb, fe_summary)
    else:
        print(f'  [WARN] Frontend docs not found at: {FRONTEND_DIR}')

    wb.save(OUT_FILE)
    be_sheets = 1
    fe_sheets = len(fe_summary) + (1 if fe_summary else 0)
    print(f'\n✅  Saved  →  {OUT_FILE}')
    print(f'   Total endpoints  : {glob_idx - 1}')
    print(f'   Backend modules  : {len(modules)}')
    print(f'   Frontend groups  : {len(fe_summary)}')
    print(f'   Total sheets     : {be_sheets + fe_sheets}')


if __name__ == '__main__':
    build_excel()
