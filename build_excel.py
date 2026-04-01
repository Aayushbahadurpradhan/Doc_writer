"""
build_excel.py
──────────────────────────────────────────────────────────────────────────────
Generates TWO colour-coded Excel workbooks from generated .md documentation.

  1. nuerabenefits_backend.xlsx
       Backend API detail — columns matched to Book.xlsx reference layout.

  2. nuerabenefits_frontend.xlsx
       Frontend page detail — columns matched to RE-Frontend Detail - Copy.xlsx
       reference layout.

Missing / incomplete rows are collected into a "⚠ Needs Fill" sheet in each
workbook so devs know exactly what still needs manual entry or AI generation.
"""

import argparse
import os
import re
import sys
from datetime import datetime

# Reconfigure stdout to UTF-8 so emoji in sheet-names don't crash on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────── CONFIG / PATHS ───────────────────────────────
# When running as a standalone script, these defaults point to the first
# sub-folder found inside doc_output/ next to this file.
# When called via  generate_excel()  (from main.py --excel), all values are
# overridden by the caller — do NOT change these constants directly.

def _default_project() -> str:
    doc_output = os.path.join(os.path.dirname(os.path.abspath(__file__)), "doc_output")
    if os.path.isdir(doc_output):
        dirs = sorted(d for d in os.listdir(doc_output)
                      if os.path.isdir(os.path.join(doc_output, d)))
        if dirs:
            return dirs[0]
    return "project"

PROJECT_NAME = _default_project()
OUTPUT_DIR      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "doc_output", PROJECT_NAME)
BACKEND_PATH    = os.path.join(OUTPUT_DIR, "docs", "backend")
FRONTEND_PATH   = os.path.join(OUTPUT_DIR, "docs", "frontend")
BACKEND_OUTPUT  = os.path.join(OUTPUT_DIR, "{}_backend.xlsx".format(PROJECT_NAME))
FRONTEND_OUTPUT = os.path.join(OUTPUT_DIR, "{}_frontend.xlsx".format(PROJECT_NAME))

# ──────────────────────────────── COLOUR PALETTE ──────────────────────────────
C_TITLE_BG       = "1B2A4A"
C_META_LABEL_BG  = "2563EB"
C_META_VALUE_BG  = "2D3A58"
C_HEADER_BG      = "243B6E"
C_ROW_ODD        = "EFF6FF"
C_ROW_EVEN       = "FFFFFF"
C_HTTP_GET       = "DCFCE7"
C_HTTP_POST      = "EDE9FE"
C_HTTP_PUT       = "FEF3C7"
C_HTTP_PATCH     = "FEF9C3"
C_HTTP_DELETE    = "FFE4E6"
C_RESPONSE       = "FAFAFA"
C_DB_ODD         = "E8F0FE"
C_DB_EVEN        = "F1F5F9"
C_STATUS         = "F0F4FF"
C_ANSWER         = "F0FDF4"
C_MISSING_ROW    = "FFF9C4"
C_MISSING_HDR    = "F59E0B"
C_TEXT_WHITE     = "FFFFFF"
C_TEXT_DARK      = "1E293B"
C_TEXT_BLUE      = "1D4ED8"
C_TEXT_AMBER     = "92400E"
C_TEXT_GREY      = "64748B"

# Frontend teal palette
C_FE_TITLE_BG    = "134E4A"
C_FE_HEADER_BG   = "0F766E"
C_FE_META_LABEL  = "0D9488"
C_FE_ROW_ODD     = "F0FDFA"
C_FE_ROW_EVEN    = "FFFFFF"

HTTP_COLOR = {
    "GET":    C_HTTP_GET,
    "POST":   C_HTTP_POST,
    "PUT":    C_HTTP_PUT,
    "PATCH":  C_HTTP_PATCH,
    "DELETE": C_HTTP_DELETE,
}

# ─────────────────────────────── STYLE HELPERS ────────────────────────────────

def _safe_title(name):
    """Strip variation selectors and enforce Excel's 31-char sheet-name limit."""
    name = name.replace("\ufe0f", "")  # remove emoji variation selector-16
    return name[:31]

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=C_TEXT_DARK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Segoe UI")

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style="thin"):
    s = Side(style=style, color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _make_writer(ws, row_num, default_bg):
    """Return a cell-writing helper bound to the given row and background."""
    def _c(col, value, bg=None, bold=False, italic=False, h="left", color=C_TEXT_DARK):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.font      = _font(bold=bold, italic=italic, color=color, size=9)
        cell.fill      = _fill(bg if bg is not None else default_bg)
        cell.alignment = _align(h=h, wrap=True)
        cell.border    = _border()
        return cell
    return _c

# ─────────────────────────────── MARKDOWN PARSERS ─────────────────────────────

def _text_between(text, start_marker, end_markers):
    m = re.search(re.escape(start_marker), text, re.IGNORECASE)
    if not m:
        return ""
    tail    = text[m.end():]
    end_pos = len(tail)
    for em in end_markers:
        ep = re.search(re.escape(em), tail, re.IGNORECASE)
        if ep and ep.start() < end_pos:
            end_pos = ep.start()
    return tail[:end_pos].strip()

def _md_table_to_text(md_table):
    rows = []
    for line in md_table.strip().splitlines():
        line = line.strip()
        if not line or re.match(r"^\|[-| :]+\|$", line):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        rows.append("  |  ".join(c for c in cells if c))
    return "\n".join(rows)

def _clean(t):
    t = re.sub(r"^\s*[—-]\s*None\s*$", "", t, flags=re.MULTILINE)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()

def parse_business_md(path):
    if not os.path.exists(path):
        return []
    text    = open(path, encoding="utf-8").read()
    results = []
    for sec in re.split(r"\n(?=## )", text):
        sec = sec.strip()
        if not sec or sec.startswith("# Business Logic"):
            continue
        name_m = re.match(r"^## (.+)", sec)
        name   = name_m.group(1).strip() if name_m else "Unknown"

        ep_m   = re.search(r"\*\*Endpoint\*\*\s*\|\s*`([^`]+)`",   sec)
        ctrl_m = re.search(r"\*\*Controller\*\*\s*\|\s*`([^`]+)`", sec)
        auth_m = re.search(r"\*\*Auth Required\*\*\s*\|\s*([^\n|]+)", sec)
        http_m = re.search(r"\*\*HTTP Method\*\*\s*\|\s*([A-Z]+)", sec)

        if not ep_m:
            fb       = re.search(r"`((?:GET|POST|PUT|DELETE|PATCH)\s+/[^`]+)`", sec)
            endpoint = fb.group(1).strip() if fb else ""
        else:
            endpoint = ep_m.group(1).strip()

        if not http_m:
            fb2         = re.search(r"\*\*HTTP\*\*:\s*([A-Z]+)", sec)
            http_method = fb2.group(1).strip() if fb2 else ""
        else:
            http_method = http_m.group(1).strip()

        if not http_method and endpoint:
            parts = endpoint.split()
            if parts and parts[0] in ("GET", "POST", "PUT", "PATCH", "DELETE"):
                http_method = parts[0]
                endpoint    = " ".join(parts[1:])
        # Also strip verb if http_method was already found but endpoint still has it
        if http_method and endpoint:
            parts = endpoint.split()
            if parts and parts[0] in ("GET", "POST", "PUT", "PATCH", "DELETE"):
                endpoint = " ".join(parts[1:]).strip()

        controller = ctrl_m.group(1).strip() if ctrl_m else ""
        auth_raw   = auth_m.group(1).strip() if auth_m else ""
        if not auth_raw:
            am2      = re.search(r"\*\*Auth\*\*:\s*`([^`]+)`", sec)
            auth_raw = am2.group(1).strip() if am2 else ""

        markers = ["### Purpose","### Business Logic","### Input Parameters",
                   "### Database Operations","### Side Effects","---","## "]
        purpose        = _text_between(sec, "### Purpose",             markers[1:])
        business_logic = _text_between(sec, "### Business Logic",      markers[2:])
        input_raw      = _text_between(sec, "### Input Parameters",    markers[3:])
        db_raw         = _text_between(sec, "### Database Operations", markers[4:])
        side_raw       = _text_between(sec, "### Side Effects",        ["---","## "])

        input_params = _md_table_to_text(input_raw) if "|" in input_raw else _clean(input_raw)
        if input_params.lower() in ("none","— none",""):
            input_params = "—"

        # ── Parse side effects into structured fields ──────────────────────
        def _se(label):
            m = re.search(rf"\*\*{re.escape(label)}\*\*:\s*([^\n]+)", side_raw)
            v = m.group(1).strip() if m else ""
            if v.lower() in ("none","- none","— none",""):
                return "None"
            return v

        side_effects = _clean(side_raw) or "None"
        emails       = _se("Emails")
        jobs         = _se("Jobs/Queues")
        events       = _se("Events")
        ext_apis     = _se("External APIs")
        files        = _se("Files")
        # Build combined readable side effects string
        se_lines = []
        if emails     != "None": se_lines.append(f"Emails: {emails}")
        if jobs        != "None": se_lines.append(f"Jobs: {jobs}")
        if events      != "None": se_lines.append(f"Events: {events}")
        if ext_apis    != "None": se_lines.append(f"Ext APIs: {ext_apis}")
        if files       != "None": se_lines.append(f"Files: {files}")
        side_effects_full = "\n".join(se_lines) if se_lines else "None"

        results.append({
            "name":              name,
            "endpoint":          endpoint,
            "http_method":       http_method.upper() if http_method else "—",
            "controller":        controller,
            "auth":              auth_raw or "verify.internal.token",
            "purpose":           _clean(purpose) or "—",
            "business_logic":    _clean(business_logic) or "—",
            "input_params":      input_params,
            "db_ops":            _clean(db_raw) or "—",
            "side_effects":      side_effects_full,
            # raw structured side-effect sub-fields (for merging)
            "_emails":   emails,
            "_jobs":     jobs,
            "_events":   events,
            "_ext_apis": ext_apis,
            "_files":    files,
        })

    # ── Bug fix #4: deduplicate endpoints — same http_method+endpoint keeps first ──
    seen: set = set()
    deduped = []
    for ep in results:
        key = (ep["http_method"], ep["endpoint"])
        if key not in seen:
            seen.add(key)
            deduped.append(ep)
    return deduped


def parse_api_md(path):
    """
    Parse api.md → dict: endpoint_url → {middleware, params}.
    api.md lists: **Endpoint**, **Controller**, **Middleware**, **Params**
    """
    if not os.path.exists(path):
        return {}
    text = open(path, encoding="utf-8").read()
    out  = {}
    for sec in re.split(r"\n(?=## )", text):
        sec = sec.strip()
        if not sec or sec.startswith("# API"):
            continue
        ep_m   = re.search(r"\*\*Endpoint\*\*\s*:\s*`([^`]+)`",    sec)
        mw_m   = re.search(r"\*\*Middleware\*\*\s*:\s*([^\n]+)",    sec)
        par_m  = re.search(r"\*\*Params\*\*\s*:\s*([^\n]+)",        sec)
        if not ep_m:
            continue
        raw_ep = ep_m.group(1).strip()
        # Strip leading HTTP verb if present: "GET /v1/..." → "/v1/..."
        parts = raw_ep.split()
        url   = parts[1] if len(parts) >= 2 and parts[0].isupper() else raw_ep
        # Bug fix #6: strip backtick wrappers from params like `{achYear}`, `{achMonth}`
        raw_params = par_m.group(1).strip() if par_m else ""
        clean_params = re.sub(r"`([^`]+)`", r"\1", raw_params)
        out[url] = {
            "middleware": mw_m.group(1).strip() if mw_m else "",
            "params":     clean_params,
        }
    return out


def parse_legacy_sql(path):
    """
    Parse legacy_query.sql → dict: endpoint_url → list of query dicts.

    Format per endpoint:
      -- ----...  (separator)
      -- Endpoint  : GET /v1/...
      -- Controller: ...
      -- ----...  (separator)
      ### N -- Query M: title
      | Field | Value | table rows |
      ```sql ... ```
    """
    if not os.path.exists(path):
        return {}
    text = open(path, encoding="utf-8").read()
    out: dict = {}

    # Split on separator lines: "-- " followed by 10+ dashes
    # Structure: EMPTY, HDR_1, BODY_1, HDR_2, BODY_2, ...
    chunks = re.split(r"--\s+-{10,}[^\n]*\n", text)

    def _qval(label, txt):
        m = re.search(
            rf"\|\s*\*\*{re.escape(label)}\*\*\s*\|\s*([^|\n]+)",
            txt, re.IGNORECASE,
        )
        v = m.group(1).strip() if m else ""
        return v if v.lower() not in ("none","—","") else ""

    # Pair every (header_chunk, body_chunk): indices (1,2), (3,4), (5,6)...
    for i in range(1, len(chunks) - 1, 2):
        hdr_chunk = chunks[i]
        body_chunk = chunks[i + 1] if i + 1 < len(chunks) else ""

        ep_m = re.search(r"-- Endpoint\s*:\s*(?:[A-Z]+\s+)?(/[^\n]+)", hdr_chunk)
        if not ep_m:
            continue
        url = ep_m.group(1).strip()

        queries = []
        for qsec in re.split(r"(?=###)", body_chunk):
            qsec = qsec.strip()
            if not qsec.startswith("###"):
                continue

            raw_sql_m = re.search(r"```sql\s*(.*?)```", qsec, re.DOTALL)
            raw_sql   = raw_sql_m.group(1).strip() if raw_sql_m else ""

            q = {
                "query_type":      _qval("Type",           qsec),
                "operation":       _qval("Operation",      qsec),
                "tables":          _qval("Tables",         qsec),
                "columns_read":    _qval("Columns Read",   qsec),
                "columns_written": _qval("Columns Written",qsec),
                "conditions":      _qval("Conditions",     qsec),
                "joins":           _qval("Joins",          qsec),
                "order_group":     _qval("Order / Group",  qsec),
                "aggregates":      _qval("Aggregates",     qsec),
                "transaction":     _qval("Transaction",    qsec),
                "soft_deletes":    _qval("Soft Deletes",   qsec),
                "raw_sql":         raw_sql,
            }
            if any(q.values()):
                queries.append(q)

        if queries:
            if url not in out:
                out[url] = []
            out[url].extend(queries)

    return out


def _sql_queries_to_text(queries):
    """Format a list of query dicts into a readable multi-line cell value."""
    if not queries:
        return "—"
    lines = []
    for i, q in enumerate(queries, 1):
        prefix = f"Query {i}: " if len(queries) > 1 else ""
        op     = q.get("operation","")
        tables = q.get("tables","")
        qtype  = q.get("query_type","")
        if op and tables:
            lines.append(f"{prefix}{op} {tables}" + (f" [{qtype}]" if qtype else ""))
        elif op:
            lines.append(f"{prefix}{op}")
    return "\n".join(lines) if lines else "—"


def _sql_conditions_to_text(queries):
    """Format conditions + joins for all queries."""
    if not queries:
        return "—"
    lines = []
    for i, q in enumerate(queries, 1):
        prefix = f"Q{i}: " if len(queries) > 1 else ""
        cond   = q.get("conditions","")
        joins  = q.get("joins","")
        if cond:
            lines.append(f"{prefix}WHERE {cond}")
        if joins:
            lines.append(f"{prefix}JOIN: {joins}")
    return "\n".join(lines) if lines else "—"


def _sql_details_to_text(queries):
    """Format columns, aggregates, transaction info."""
    if not queries:
        return "—"
    lines = []
    for i, q in enumerate(queries, 1):
        prefix = f"Q{i}: " if len(queries) > 1 else ""
        cols_r  = q.get("columns_read","")
        cols_w  = q.get("columns_written","")
        agg     = q.get("aggregates","")
        order   = q.get("order_group","")
        txn     = q.get("transaction","")
        soft    = q.get("soft_deletes","")
        raw     = q.get("raw_sql","")
        if cols_r and cols_r != "*":
            lines.append(f"{prefix}Read: {cols_r}")
        if cols_w:
            lines.append(f"{prefix}Write: {cols_w}")
        if agg:
            lines.append(f"{prefix}Aggregates: {agg}")
        if order:
            lines.append(f"{prefix}Order/Group: {order}")
        if txn:
            lines.append(f"{prefix}Txn: {txn}")
        if soft:
            lines.append(f"{prefix}SoftDel: {soft}")
        if raw:
            lines.append(f"{prefix}SQL: {raw[:200]}")
    return "\n".join(lines) if lines else "—"


def parse_responses_md(path):
    """
    Parse responses.md → dict: endpoint_url → rich response dict.

    Extracts per endpoint:
      response_type, path_params, fields_json, example_json, description
    """
    if not os.path.exists(path):
        return {}
    text = open(path, encoding="utf-8").read()
    out  = {}
    for sec in re.split(r"\n(?=## )", text):
        sec = sec.strip()
        if not sec or sec.startswith("# API Response"):
            continue
        hdr_m = re.match(r"^## (.+)", sec)
        if not hdr_m:
            continue
        hdr = hdr_m.group(1).strip()

        # Extract the URL part: "GET /v1/view-ach/{achYear}/{achMonth}" → "/v1/..."
        url_m = re.match(r"^(?:[A-Z]+\s+)?(/\S+)", hdr)
        url   = url_m.group(1).strip() if url_m else hdr

        # Response type
        rt_m  = re.search(r"\*\*Response Type\*\*\s*:\s*`?([^\n`]+)`?", sec)
        rtype = rt_m.group(1).strip() if rt_m else ""

        # Path parameters
        pp_m    = re.search(
            r"\*\*Path Parameters\*\*[:\s]*(.*?)(?=\*\*|\n\n|\Z)", sec, re.DOTALL
        )
        path_params = ""
        if pp_m:
            pp_block = pp_m.group(1).strip()
            pp_list  = [ln.lstrip("- ").strip() for ln in pp_block.splitlines()
                        if ln.strip().startswith("-")]
            path_params = "; ".join(pp_list) if pp_list else pp_block[:120]

        # Response fields JSON (first ```json block)
        json_blocks = re.findall(r"```json\s*(.*?)```", sec, re.DOTALL)
        fields_json  = json_blocks[0].strip() if json_blocks else ""
        example_json = json_blocks[1].strip() if len(json_blocks) > 1 else ""

        # Description
        desc_m  = re.search(
            r"\*\*Description\*\*[:\s]*(.*?)(?=\n\n|\n---|\Z)", sec, re.DOTALL
        )
        description = desc_m.group(1).strip() if desc_m else ""

        out[url] = {
            "response_type":  rtype,
            "path_params":    path_params,
            "fields_json":    fields_json,
            "example_json":   example_json,
            "description":    description,
            "_raw_hdr":       hdr,
        }
    return out


def _url_fuzzy_match(url: str, mapping: dict):
    """
    Bug fix #2: safe fuzzy match that uses path segments, not substring.

    Returns the matching value from `mapping`, or an empty dict / empty list
    depending on what values the mapping holds.

    Rules (in priority order):
      1. Exact match (already tried by caller)
      2. Match by normalised path (collapse numeric IDs and path params to {id})
      3. Match by path suffix (last 2 segments)
    Never matches /bill → /billing (segmented, not substring).
    """
    empty = [] if mapping and isinstance(next(iter(mapping.values())), list) else {}

    if not url or not mapping:
        return empty

    def _norm(u):
        # collapse /{numeric} and /{word_param} to /{id}
        u = re.sub(r"/\d+", "/{id}", u.lower())
        u = re.sub(r"/\{[^}]+\}", "/{id}", u)
        return u.rstrip("/")

    norm_url = _norm(url)
    url_segs = [s for s in norm_url.split("/") if s]

    # Build normalised index
    for candidate_url, val in mapping.items():
        norm_cand = _norm(candidate_url)
        if norm_cand == norm_url:
            return val
        # Match by last-2 segments (same endpoint, different version prefix)
        cand_segs = [s for s in norm_cand.split("/") if s]
        if len(url_segs) >= 2 and len(cand_segs) >= 2:
            if url_segs[-2:] == cand_segs[-2:]:
                return val

    return empty


def load_domain(domain_path):
    """
    Load ALL data for a domain by merging all 4 source files:
      business.md, api.md, legacy_query.sql, responses.md
    """
    endpoints  = parse_business_md(  os.path.join(domain_path, "business.md"))
    api_info   = parse_api_md(        os.path.join(domain_path, "api.md"))
    sql_info   = parse_legacy_sql(    os.path.join(domain_path, "legacy_query.sql"))
    resp_map   = parse_responses_md(  os.path.join(domain_path, "responses.md"))

    for ep in endpoints:
        url = ep["endpoint"]

        # ── Merge api.md (middleware, path params) ────────────────────────────
        # Bug fix #2: use path-segment safe fuzzy match to avoid /bill matching /billing
        api_entry = api_info.get(url, {})
        if not api_entry:
            api_entry = _url_fuzzy_match(url, api_info)
        ep["middleware"]  = api_entry.get("middleware", ep.get("auth","verify.internal.token"))
        ep["route_params"]= api_entry.get("params", "")

        # ── Merge legacy_query.sql ────────────────────────────────────────────
        sql_queries = sql_info.get(url, [])
        if not sql_queries:
            sql_queries = _url_fuzzy_match(url, sql_info)  # returns [] if not found
        ep["sql_queries"]       = sql_queries
        ep["db_op_tables"]      = _sql_queries_to_text(sql_queries)
        ep["db_conditions"]     = _sql_conditions_to_text(sql_queries)
        ep["db_details"]        = _sql_details_to_text(sql_queries)
        # Fallback: keep old db_ops if sql_queries is empty
        if ep["db_op_tables"] == "—" and ep.get("db_ops","—") != "—":
            ep["db_op_tables"] = ep["db_ops"]

        # ── Merge responses.md ────────────────────────────────────────────────
        # Bug fix #3: removed positional fallback — it assigns wrong responses
        resp_entry = resp_map.get(url, {})
        if not resp_entry:
            matched_resp = _url_fuzzy_match(url, resp_map)
            resp_entry = matched_resp if isinstance(matched_resp, dict) else {}

        ep["response_type"]   = resp_entry.get("response_type","")   or "json"
        ep["response_fields"] = resp_entry.get("fields_json","")     or "—"
        ep["response_example"]= resp_entry.get("example_json","")    or "—"
        ep["response_desc"]   = resp_entry.get("description","")     or "—"
        ep["response"]        = ep["response_fields"]   # backward compat

        # ── Bug fix #5: extract External/Internal API calls as own field ──────
        ext_api_full = ep.get("_ext_apis", "")
        # Also scan business_logic text for internal service/API call patterns
        bl = ep.get("business_logic", "")
        internal_calls = []
        for line in bl.splitlines():
            ll = line.lower()
            if any(kw in ll for kw in (
                "calls another", "calls the", "calls an",
                "internally calls", "uses the", "uses a",
                "service class", "->service", "->repository", "->repo",
                "fire()", "dispatch(", "queue(", "job(",
                "sends an email", "notif",
            )):
                internal_calls.append(line.strip().lstrip("- ").strip())
        combined_api_calls = ""
        if ext_api_full and ext_api_full.lower() not in ("none", ""):
            combined_api_calls += ext_api_full
        if internal_calls:
            if combined_api_calls:
                combined_api_calls += "\n"
            combined_api_calls += "\n".join(internal_calls)
        ep["ext_internal_calls"] = combined_api_calls or "None"

    return endpoints

# ──────────────────────────── MISSING DATA HELPERS ────────────────────────────

_REQUIRED_BE = ["name","endpoint","http_method","purpose","controller"]

def _find_missing(ep):
    return [f for f in _REQUIRED_BE if not ep.get(f) or ep[f] in ("—","Unknown","")]

def _be_status_codes(http):
    return {
        "GET":    "200 OK\n404 Not Found\n422 Validation Error",
        "POST":   "201 Created\n200 OK\n422 Validation Error\n404 Not Found",
        "PUT":    "200 OK\n422 Validation Error\n404 Not Found",
        "PATCH":  "200 OK\n422 Validation Error\n404 Not Found",
        "DELETE": "200 OK\n404 Not Found",
    }.get(http.upper(), "200 OK\n422 Validation Error\n404 Not Found")

def _build_missing_sheet(wb, missing_log, today_str, mode="backend"):
    title = "⚠ Missing Backend" if mode == "backend" else "⚠ Missing Frontend"
    ws    = wb.create_sheet(title=_safe_title(title))
    ws.freeze_panes = "A4"
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    c = ws.cell(row=1, column=1,
                value="⚠  Incomplete / Missing Data — Needs Manual Fill or AI Generation")
    c.font      = _font(bold=True, color=C_TEXT_WHITE, size=13)
    c.fill      = _fill("92400E")
    c.alignment = _align(h="left", wrap=False)
    ws.row_dimensions[2].height = 18
    ws.cell(row=2, column=1, value=f"Generated: {today_str}").font = _font(size=9, italic=True)
    ws.cell(row=2, column=2,
            value=f"Total items needing attention: {len(missing_log)}").font = _font(size=9, italic=True)
    headers = ["#","Domain / Group","Endpoint / Screen","Missing Fields","Action Required"]
    widths  = [4, 22, 46, 42, 28]
    ws.row_dimensions[3].height = 28
    for idx, (hdr, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=3, column=idx, value=hdr)
        cell.font      = _font(bold=True, color=C_TEXT_WHITE, size=10)
        cell.fill      = _fill(C_MISSING_HDR)
        cell.alignment = _align(h="center", v="center")
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(idx)].width = w
    ai_fields = {"purpose","response","db_ops","business_logic","api_calls","http_method"}
    for i, item in enumerate(missing_log, start=1):
        row        = i + 3
        bg         = C_MISSING_ROW if (i % 2 == 1) else "FFFDE7"
        missing_f  = item.get("missing","")
        name_ep    = item.get("endpoint") or item.get("screen") or item.get("name","?")
        domain_grp = item.get("domain") or item.get("group","?")
        is_ai      = any(f.strip() in ai_fields for f in missing_f.split(","))
        action     = "AI Fill Recommended" if is_ai else "Manual Fill"
        ws.row_dimensions[row].height = 20
        _c = _make_writer(ws, row, bg)
        _c(1, i,          h="center")
        _c(2, domain_grp)
        _c(3, name_ep)
        _c(4, missing_f,  color=C_TEXT_AMBER, bold=True)
        _c(5, action,     color="15803D" if "AI" in action else C_TEXT_AMBER)
    return ws

# ══════════════════════════════════════════════════════════════════════════════
#  BACKEND WORKBOOK  — columns matched to Book.xlsx
# ══════════════════════════════════════════════════════════════════════════════

BE_COLUMNS = [
    # ── Identity ─────────────────────────────────────────
    ("#",                               4),
    ("Function /\nEndpoint Name",      22),
    ("Purpose",                        34),
    ("HTTP\nMethod",                    9),
    ("Endpoint URL",                   36),
    # ── Request ──────────────────────────────────────────
    ("Route\nParameters",              22),   # path params from api.md / responses.md
    ("Request Body\n/ Query Params",   34),   # body / query params from business.md
    # ── Business ─────────────────────────────────────────
    ("Business Rules\n/ Notes",        34),
    # ── Database ─────────────────────────────────────────
    ("DB Operation\n& Tables",         26),   # operation + table names from legacy_query.sql
    ("DB Conditions\n& Joins",         28),   # WHERE conditions + JOINs
    ("DB Query\nDetails",              26),   # columns, aggregates, transaction, raw SQL
    # ── Side Effects ─────────────────────────────────────
    ("Side Effects",                   32),   # emails, jobs, events, ext APIs, files
    ("External / Internal\nAPI Calls",  38),   # ext APIs + internal service/cross-API calls
    # ── Q&A ──────────────────────────────────────────────
    ("Open Questions",                 24),
    ("Answer / Decision",              24),
    ("Answered By",                    14),
    ("Date Answered",                  14),
    # ── Response ─────────────────────────────────────────
    ("Response\nType",                 14),   # json / array_of_objects
    ("Response Schema\n(Fields)",      38),   # JSON schema of response
    ("Example Response",               38),   # example JSON payload
    ("Response\nDescription",          36),   # description text
    # ── Status / Auth ─────────────────────────────────────
    ("Status Codes",                   16),
    ("Auth /\nMiddleware",             22),
    # ── Meta ─────────────────────────────────────────────
    ("Priority",                       10),
    ("Owner\n(Controller)",            26),
    ("Last Updated",                   14),
]
BE_NUM_COLS = len(BE_COLUMNS)

def _write_be_header(ws):
    ws.row_dimensions[3].height = 36
    for col_idx, (col_name, _) in enumerate(BE_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font      = _font(bold=True, color=C_TEXT_WHITE, size=10)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = _align(h="center", v="center", wrap=True)
        cell.border    = _border()

def _write_title_meta(ws, title_text, meta_pairs, num_cols,
                      title_bg=C_TITLE_BG, meta_label_bg=C_META_LABEL_BG,
                      meta_value_bg=C_META_VALUE_BG):
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    t = ws.cell(row=1, column=1, value=title_text)
    t.font      = _font(bold=True, color=C_TEXT_WHITE, size=13)
    t.fill      = _fill(title_bg)
    t.alignment = _align(h="left", wrap=False)
    ws.row_dimensions[2].height = 20
    col = 1
    for label, value in meta_pairs:
        lc = ws.cell(row=2, column=col, value=label)
        lc.font      = _font(bold=True, color=C_TEXT_WHITE, size=10)
        lc.fill      = _fill(meta_label_bg)
        lc.alignment = _align(h="right", wrap=False)
        vc = ws.cell(row=2, column=col + 1, value=value)
        vc.font      = _font(color=C_TEXT_WHITE, size=10)
        vc.fill      = _fill(meta_value_bg)
        vc.alignment = _align(h="left", wrap=False)
        col += 2
    for c in range(col, num_cols + 1):
        ws.cell(row=2, column=c).fill = _fill(meta_value_bg)

def build_backend_workbook(domains_data, domain_summary, today_str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    missing_log = []

    # ── Index sheet ────────────────────────────────────────────────────────
    ws_idx = wb.create_sheet(title=_safe_title("📋 Index"))
    ws_idx.freeze_panes = "A4"
    _write_title_meta(
        ws_idx,
        f"📋  {PROJECT_NAME.upper()} — Backend API Documentation Index",
        [("Project:", PROJECT_NAME), ("Generated:", today_str),
         ("Total Domains:", str(len(domain_summary)))],
        6,
    )
    idx_headers = ["#","Domain","Endpoints","Controller(s)","Version / Prefix","Description"]
    idx_widths  = [4,   20,     12,          38,              18,                52]
    ws_idx.row_dimensions[3].height = 28
    for col_idx, (hdr, w) in enumerate(zip(idx_headers, idx_widths), start=1):
        cell = ws_idx.cell(row=3, column=col_idx, value=hdr)
        cell.font      = _font(bold=True, color=C_TEXT_WHITE, size=10)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = _align(h="center", v="center")
        cell.border    = _border()
        ws_idx.column_dimensions[get_column_letter(col_idx)].width = w
    for i, (domain, ep_count, controllers, version, description) in \
            enumerate(domain_summary, start=1):
        row = i + 3
        bg  = C_ROW_ODD if (i % 2 == 1) else C_ROW_EVEN
        ws_idx.row_dimensions[row].height = 18
        _c = _make_writer(ws_idx, row, bg)
        _c(1, i,            h="center")
        _c(2, domain,       bold=True, color=C_TEXT_BLUE)
        _c(3, ep_count,     h="center")
        _c(4, controllers)
        _c(5, version)
        _c(6, description)

    # ── Combined backend sheet ─────────────────────────────────────────────
    ws = wb.create_sheet(title=_safe_title(f"{PROJECT_NAME} Backend"))
    ws.freeze_panes = "A4"
    for col_idx, (_, width) in enumerate(BE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    _write_title_meta(
        ws,
        f"🔌  API Detail Sheet — {PROJECT_NAME.upper()} BACKEND",
        [("Module:", PROJECT_NAME), ("Last Updated:", today_str),
         ("Status:", "In Progress"), ("API Version:", "V1 / V2 / V3")],
        BE_NUM_COLS,
    )
    _write_be_header(ws)

    row_num  = 4
    glob_idx = 1

    for domain_name, endpoints in domains_data:
        if not endpoints:
            continue
        # Domain divider row
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=BE_NUM_COLS)
        s = ws.cell(row=row_num, column=1, value=f"  ◆  {domain_name.upper()}")
        s.fill      = _fill(C_TITLE_BG)
        s.font      = _font(bold=True, color=C_TEXT_WHITE, size=10)
        s.alignment = _align(h="left", wrap=False)
        ws.row_dimensions[row_num].height = 22
        row_num += 1

        for i, ep in enumerate(endpoints):
            is_odd = (i % 2 == 0)
            row_bg = C_ROW_ODD if is_odd else C_ROW_EVEN
            db_bg  = C_DB_ODD  if is_odd else C_DB_EVEN
            http   = ep.get("http_method", "—")
            c_http = HTTP_COLOR.get(http.upper(), C_ROW_ODD)
            missing = _find_missing(ep)
            if missing:
                missing_log.append({
                    "domain":   domain_name,
                    "name":     ep.get("name","?"),
                    "endpoint": ep.get("endpoint",""),
                    "missing":  ", ".join(missing),
                })
                row_bg = C_MISSING_ROW

            # Dynamic row height: tallest multi-line field determines height
            line_counts = [
                len(str(ep.get("input_params","") or "").split("\n")),
                len(str(ep.get("db_op_tables","") or "").split("\n")),
                len(str(ep.get("db_conditions","") or "").split("\n")),
                len(str(ep.get("db_details","") or "").split("\n")),
                len(str(ep.get("business_logic","") or "").split("\n")),
                len(str(ep.get("side_effects","") or "").split("\n")),
                3,
            ]
            ws.row_dimensions[row_num].height = max(60, 15 * max(line_counts))

            _c = _make_writer(ws, row_num, row_bg)
            # ── Identity ──────────────────────────────────────────────────────
            _c(1,  glob_idx,
               bg=C_TITLE_BG, bold=True, h="center", color=C_TEXT_WHITE)
            _c(2,  ep.get("name","—"),             bold=True)
            _c(3,  ep.get("purpose","—"))
            _c(4,  http, bg=c_http,                bold=True, h="center")
            _c(5,  ep.get("endpoint","—"),         bold=True, color=C_TEXT_BLUE)
            # ── Request ───────────────────────────────────────────────────────
            _c(6,  ep.get("route_params","—") or "—",  bg=C_RESPONSE)
            _c(7,  ep.get("input_params","—"),          bg=C_RESPONSE)
            # ── Business ──────────────────────────────────────────────────────
            _c(8,  ep.get("business_logic","—"),   bg=db_bg)
            # ── Database ──────────────────────────────────────────────────────
            _c(9,  ep.get("db_op_tables","—"),     bg=db_bg)
            _c(10, ep.get("db_conditions","—"),    bg=db_bg)
            _c(11, ep.get("db_details","—"),       bg=db_bg)
            # ── Side Effects ──────────────────────────────────────────────────
            _c(12, ep.get("side_effects","None"),           bg=db_bg)
            _c(13, ep.get("ext_internal_calls","None"),     bg=db_bg)
            # ── Q&A ───────────────────────────────────────────────────────────
            _c(14, "")
            _c(15, "", bg=C_ANSWER)
            _c(16, "")
            _c(17, "")
            # ── Response ──────────────────────────────────────────────────────
            _c(18, ep.get("response_type","json"),  bg=C_STATUS, h="center")
            _c(19, ep.get("response_fields","—"),   bg=C_RESPONSE)
            _c(20, ep.get("response_example","—"),  bg=C_RESPONSE)
            _c(21, ep.get("response_desc","—"),     bg=C_RESPONSE)
            # ── Status / Auth ─────────────────────────────────────────────────
            _c(22, _be_status_codes(http),          bg=C_STATUS)
            _c(23, ep.get("middleware","—") or ep.get("auth","—"))
            # ── Meta ──────────────────────────────────────────────────────────
            _c(24, "—", bg=C_ROW_ODD, h="center")
            _c(25, ep.get("controller","—"),        color=C_TEXT_GREY)
            _c(26, today_str,                       h="center")
            row_num  += 1
            glob_idx += 1

    ws.auto_filter.ref = f"A3:{get_column_letter(BE_NUM_COLS)}{row_num - 1}"
    if missing_log:
        _build_missing_sheet(wb, missing_log, today_str, "backend")
    return wb, glob_idx - 1, missing_log

# ══════════════════════════════════════════════════════════════════════════════
#  FRONTEND WORKBOOK  — columns matched to RE-Frontend Detail - Copy.xlsx
# ══════════════════════════════════════════════════════════════════════════════

FE_COLUMNS = [
    ("#",                                   4),
    ("Screen Name",                        26),
    ("Route / URL",                        30),
    ("Vue Component Path",                 38),
    ("API Endpoint",                       36),
    ("HTTP Method",                        12),
    ("Request Payload /\nQuery Parameters", 38),
    ("Conditional Logic",                  32),
    ("Validation Rules",                   32),
]
FE_NUM_COLS = len(FE_COLUMNS)

def _write_fe_header(ws):
    ws.row_dimensions[3].height = 36
    for col_idx, (col_name, _) in enumerate(FE_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font      = _font(bold=True, color=C_TEXT_WHITE, size=10)
        cell.fill      = _fill(C_FE_HEADER_BG)
        cell.alignment = _align(h="center", v="center", wrap=True)
        cell.border    = _border()

def parse_frontend_md(path):
    if not os.path.exists(path):
        return []
    text  = open(path, encoding="utf-8").read()
    pages = []
    for sec in re.split(r"\n(?=# (?:Page|`/))", text):
        sec = sec.strip()
        if not sec:
            continue
        route = ""
        hm = re.match(r"^#\s+(?:Page:\s*)?`([^`]+)`", sec)
        if hm:
            route = hm.group(1).strip()

        def _field(label):
            m = re.search(
                rf"\|\s*\*\*{re.escape(label)}\*\*\s*\|\s*`?([^|\n`]+)`?\s*\|", sec)
            return m.group(1).strip() if m else ""

        component   = _field("Component")
        source_file = _field("Source file")
        layout      = _field("Layout")
        example_url = _field("Example URL")

        children = []
        cm = re.search(r"## Child Components\n+(.*?)(?=\n##|\Z)", sec, re.DOTALL)
        if cm:
            block = cm.group(1).strip()
            NONE_V = {"none","none detected",
                      "none — no imported or template sub-components detected"}
            if block.lower() not in NONE_V:
                children = [ln.lstrip("- ").strip().strip("`")
                            for ln in block.splitlines()
                            if ln.strip().startswith("-")]

        composables = []
        cmp_m = re.search(r"## Composables Used\n+(.*?)(?=\n##|\Z)", sec, re.DOTALL)
        if cmp_m:
            block = cmp_m.group(1).strip()
            if block.lower() not in ("none","none detected"):
                composables = [ln.lstrip("- ").strip().strip("`").rstrip("()")
                               for ln in block.splitlines()
                               if ln.strip().startswith("-")]

        api_calls  = []
        no_api_flag = False   # True when md explicitly says no calls detected
        api_m = re.search(r"## Backend API Dependencies\n+(.*?)(?=\n##|\Z)", sec, re.DOTALL)
        if api_m:
            block = api_m.group(1).strip()
            # Detect explicit "none detected" markers (italicised or plain)
            _block_plain = block.strip().lstrip("_").rstrip("_").lower()
            _none_phrases = (
                "none", "none detected", "none — no axios/fetch/form calls detected",
                "no axios/fetch/form calls detected", "no api calls",
                "no api calls detected",
            )
            if _block_plain in _none_phrases or _block_plain.startswith("none —"):
                no_api_flag = True
            else:
                for row in re.finditer(
                    r"\|\s*`([A-Z]+)`\s*\|\s*`([^`]+)`\s*\|([^|]+)\|([^|]+)\|", block
                ):
                    api_calls.append({
                        "method":   row.group(1).strip(),
                        "endpoint": row.group(2).strip(),
                        "source":   row.group(3).strip(),
                        "via":      row.group(4).strip(),
                    })
                if not api_calls:
                    for ep_line in re.finditer(
                        r"-\s*Endpoint:\s*`([^`]+)`.*?Method:\s*([A-Z]+)",
                        block, re.DOTALL,
                    ):
                        api_calls.append({
                            "method":   ep_line.group(2).strip(),
                            "endpoint": ep_line.group(1).strip(),
                            "source":   "",
                            "via":      "axios",
                        })

        st_m = re.search(r"## State Management\n+(.*?)(?=\n##|\Z)", sec, re.DOTALL)
        state_mgmt = st_m.group(1).strip() if st_m else "—"
        if state_mgmt.lower() in ("none","none detected","—"):
            state_mgmt = "—"

        wn_m = re.search(r"## Warnings\n+(.*?)(?=\n##|\Z)", sec, re.DOTALL)
        warnings = wn_m.group(1).strip() if wn_m else ""
        if warnings.lower() in ("none","_none_"):
            warnings = ""

        rp_m = re.search(r"## Request Payload.*?\n+(.*?)(?=\n##|\Z)", sec, re.DOTALL)
        request_payload = rp_m.group(1).strip() if rp_m else ""
        if request_payload.lower().startswith("_"):
            request_payload = ""

        cl_m = re.search(r"## Conditional Logic.*?\n+(.*?)(?=\n##|\Z)", sec, re.DOTALL)
        conditional_logic = cl_m.group(1).strip() if cl_m else ""
        if conditional_logic.lower().startswith("_"):
            conditional_logic = ""

        vr_m = re.search(r"## Validation Rules.*?\n+(.*?)(?=\n##|\Z)", sec, re.DOTALL)
        validation_rules = vr_m.group(1).strip() if vr_m else ""
        if validation_rules.lower().startswith("_"):
            validation_rules = ""

        pages.append({
            "route":             route or "UNKNOWN",
            "component":         component,
            "source_file":       source_file,
            "layout":            layout,
            "example_url":       example_url,
            "children":          children,
            "composables":       composables,
            "api_calls":         api_calls,
            "no_api_flag":       no_api_flag,   # page explicitly has no API calls
            "state_mgmt":        state_mgmt,
            "warnings":          warnings,
            "request_payload":   request_payload,
            "conditional_logic": conditional_logic,
            "validation_rules":  validation_rules,
        })
    return pages

def load_frontend_group(group_dir):
    pages = []
    for fname in sorted(os.listdir(group_dir)):
        if fname.lower() == "readme.md" or not fname.endswith(".md"):
            continue
        parsed = parse_frontend_md(os.path.join(group_dir, fname))
        for p in parsed:
            p["_filename"] = fname
        pages.extend(parsed)
    return pages

def build_frontend_workbook(fe_groups_data, today_str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    missing_fe = []
    fe_summary = []

    for group_name, pages in fe_groups_data:
        safe_title = _safe_title(re.sub(r"[\\/*?:\[\]]", "", f"FE_{group_name}"))
        ws = wb.create_sheet(title=safe_title)
        ws.freeze_panes = "A4"
        for col_idx, (_, w) in enumerate(FE_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        _write_title_meta(
            ws,
            f"🖥️  Frontend Pages — /{group_name}",
            [("Group:", f"/{group_name}"), ("Pages:", str(len(pages))),
             ("Last Updated:", today_str)],
            FE_NUM_COLS,
            title_bg=C_FE_TITLE_BG,
            meta_label_bg=C_FE_META_LABEL,
            meta_value_bg=C_META_VALUE_BG,
        )
        _write_fe_header(ws)

        if not pages:
            ws.row_dimensions[4].height = 24
            cell = ws.cell(row=4, column=1, value="No page data found for this group.")
            cell.font      = _font(italic=True, color="94A3B8", size=9)
            cell.fill      = _fill(C_FE_ROW_ODD)
            cell.alignment = _align(h="center")
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=FE_NUM_COLS)
            fe_summary.append((group_name, 0, 0))
            continue

        row_num   = 4
        total_api = 0
        for page in pages:
            api_calls         = page.get("api_calls", [])
            no_api_flag       = page.get("no_api_flag", False)
            screen_name       = page.get("component","") or page.get("route","UNKNOWN")
            route             = page.get("route","UNKNOWN")
            source_file       = page.get("source_file","—") or "—"
            request_payload   = page.get("request_payload","") or ""
            conditional_logic = page.get("conditional_logic","") or ""
            validation_rules  = page.get("validation_rules","") or ""

            if api_calls:
                for j, call in enumerate(api_calls):
                    total_api += 1
                    method   = call.get("method","?")
                    endpoint = call.get("endpoint","?")
                    c_http   = HTTP_COLOR.get(method.upper(), C_FE_ROW_ODD)
                    is_odd   = (row_num % 2 == 0)
                    row_bg   = C_FE_ROW_ODD if is_odd else C_FE_ROW_EVEN
                    ws.row_dimensions[row_num].height = 36
                    _c = _make_writer(ws, row_num, row_bg)
                    _c(1,  row_num - 3,
                       bg=C_FE_TITLE_BG, bold=True, h="center", color=C_TEXT_WHITE)
                    _c(2,  screen_name, bold=(j == 0))
                    _c(3,  route,       bold=True, color=C_TEXT_BLUE)
                    _c(4,  source_file, color=C_TEXT_GREY)
                    _c(5,  endpoint,    bold=True, color=C_TEXT_BLUE)
                    _c(6,  method, bg=c_http, bold=True, h="center")
                    _c(7,  request_payload)
                    _c(8,  conditional_logic)
                    _c(9,  validation_rules)
                    row_num += 1
            elif no_api_flag:
                # Page explicitly has no API calls — not a missing-data issue
                is_odd = (row_num % 2 == 0)
                row_bg = C_FE_ROW_ODD if is_odd else C_FE_ROW_EVEN
                ws.row_dimensions[row_num].height = 28
                _c = _make_writer(ws, row_num, row_bg)
                _c(1,  row_num - 3,
                   bg=C_FE_TITLE_BG, bold=True, h="center", color=C_TEXT_WHITE)
                _c(2,  screen_name, bold=True)
                _c(3,  route, bold=True, color=C_TEXT_BLUE)
                _c(4,  source_file, color=C_TEXT_GREY)
                _c(5,  "(No API calls)", color="64748B", italic=True)
                _c(6,  "—", h="center")
                _c(7,  request_payload)
                _c(8,  conditional_logic)
                _c(9,  validation_rules)
                row_num += 1
            else:
                # Unknown — API calls could not be detected
                ws.row_dimensions[row_num].height = 36
                _c = _make_writer(ws, row_num, C_MISSING_ROW)
                _c(1,  row_num - 3,
                   bg=C_FE_TITLE_BG, bold=True, h="center", color=C_TEXT_WHITE)
                _c(2,  screen_name, bold=True)
                _c(3,  route, bold=True, color=C_TEXT_BLUE)
                _c(4,  source_file, color=C_TEXT_GREY)
                _c(5,  "[NEEDS FILL]", color=C_TEXT_AMBER, bold=True)
                _c(6,  "—")
                _c(7,  "—")
                _c(8,  "")
                _c(9,  "")
                missing_fe.append({
                    "group":   group_name,
                    "screen":  screen_name,
                    "route":   route,
                    "missing": "api_endpoint, http_method, request_payload",
                })
                row_num += 1

        ws.auto_filter.ref = f"A3:{get_column_letter(FE_NUM_COLS)}{row_num - 1}"
        fe_summary.append((group_name, len(pages), total_api))

    # Frontend index sheet
    ws_idx = wb.create_sheet(title=_safe_title("🖥 FE Index"), index=0)
    ws_idx.freeze_panes = "A4"
    _write_title_meta(
        ws_idx,
        f"🖥️  {PROJECT_NAME.upper()} Frontend — Page Groups Index",
        [("Project:", PROJECT_NAME), ("Generated:", today_str),
         ("Groups:", str(len(fe_summary)))],
        5,
        title_bg=C_FE_TITLE_BG,
        meta_label_bg=C_FE_META_LABEL,
        meta_value_bg=C_META_VALUE_BG,
    )
    idx_hdrs = ["#","Group / Prefix","Pages","API Calls","Sheet Name"]
    idx_ws   = [4,   24,              10,      12,          28]
    ws_idx.row_dimensions[3].height = 28
    for col_idx, (hdr, w) in enumerate(zip(idx_hdrs, idx_ws), start=1):
        cell = ws_idx.cell(row=3, column=col_idx, value=hdr)
        cell.font      = _font(bold=True, color=C_TEXT_WHITE, size=10)
        cell.fill      = _fill(C_FE_HEADER_BG)
        cell.alignment = _align(h="center", v="center")
        cell.border    = _border()
        ws_idx.column_dimensions[get_column_letter(col_idx)].width = w
    for i, (group, page_count, api_count) in enumerate(fe_summary, start=1):
        row        = i + 3
        bg         = C_FE_ROW_ODD if (i % 2 == 1) else C_FE_ROW_EVEN
        sheet_name = _safe_title(re.sub(r"[\\/*?:\[\]]", "", f"FE_{group}"))
        ws_idx.row_dimensions[row].height = 18
        _c = _make_writer(ws_idx, row, bg)
        _c(1, i,            h="center")
        _c(2, f"/{group}",  bold=True, color=C_TEXT_BLUE)
        _c(3, page_count,   h="center")
        _c(4, api_count,    h="center")
        _c(5, sheet_name)

    if missing_fe:
        _build_missing_sheet(wb, missing_fe, today_str, "frontend")
    return wb, missing_fe

# ──────────────────────────────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    # ── Parse CLI arguments ──────────────────────────────────────────────
    global PROJECT_NAME, BACKEND_PATH, FRONTEND_PATH, OUTPUT_DIR
    global BACKEND_OUTPUT, FRONTEND_OUTPUT
    parser = argparse.ArgumentParser(
        description="Build Excel docs from generated markdown.",
        add_help=False,  # avoid conflicts when called from generate_excel()
    )
    parser.add_argument("--docs",    help="Path to docs/ folder (contains backend/ and frontend/)")
    parser.add_argument("--project", help="Project name for titles and filenames")
    parser.add_argument("--output",  help="Folder where .xlsx files are saved")
    args, _ = parser.parse_known_args()
    if args.project:
        PROJECT_NAME = args.project
    if args.docs:
        BACKEND_PATH  = os.path.join(args.docs, "backend")
        FRONTEND_PATH = os.path.join(args.docs, "frontend")
    if args.output:
        OUTPUT_DIR = args.output
    BACKEND_OUTPUT  = os.path.join(OUTPUT_DIR, f"{PROJECT_NAME}_backend.xlsx")
    FRONTEND_OUTPUT = os.path.join(OUTPUT_DIR, f"{PROJECT_NAME}_frontend.xlsx")

    today_str = datetime.today().strftime("%Y-%m-%d")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ══ BACKEND ════════════════════════════════════════════════════════════════
    print(f"=== Building BACKEND Excel for [{PROJECT_NAME}] ===")
    if not os.path.isdir(BACKEND_PATH):
        print(f"[ERROR] Backend path not found: {BACKEND_PATH}")
        return

    domains = sorted([
        d for d in os.listdir(BACKEND_PATH)
        if os.path.isdir(os.path.join(BACKEND_PATH, d))
    ])
    print(f"Found {len(domains)} backend domains.")

    domains_data   = []
    domain_summary = []
    for domain in domains:
        domain_path = os.path.join(BACKEND_PATH, domain)
        print(f"  Processing: {domain}")
        endpoints = load_domain(domain_path)
        domain_summary.append((domain, len(endpoints), "", "", ""))
        domains_data.append((domain, endpoints))

    wb_be, total_eps, be_missing = build_backend_workbook(
        domains_data, domain_summary, today_str
    )
    wb_be.save(BACKEND_OUTPUT)
    print(f"\nSaved => {BACKEND_OUTPUT}")
    print(f"  Sheets          : {list(wb_be.sheetnames)}")
    print(f"  Total endpoints : {total_eps}")
    print(f"  Incomplete rows : {len(be_missing)}  (see sheet: ⚠ Missing Backend)")

    # ══ FRONTEND ═══════════════════════════════════════════════════════════════
    print(f"\n=== Building FRONTEND Excel for [{PROJECT_NAME}] ===")
    if not os.path.isdir(FRONTEND_PATH):
        print(f"[WARN] Frontend docs not found at: {FRONTEND_PATH}")
        print("       Skipping frontend Excel generation.")
        return

    fe_dirs = sorted([
        d for d in os.listdir(FRONTEND_PATH)
        if os.path.isdir(os.path.join(FRONTEND_PATH, d))
        and d not in ("undocumented",)
    ])
    if not fe_dirs:
        print("[INFO] Frontend folder exists but contains no group sub-folders.")
        print("       Skipping frontend Excel generation.")
        return

    print(f"Found {len(fe_dirs)} frontend groups.")
    fe_groups_data = []
    for group in fe_dirs:
        group_dir = os.path.join(FRONTEND_PATH, group)
        print(f"  FE Group: {group}")
        pages = load_frontend_group(group_dir)
        fe_groups_data.append((group, pages))

    wb_fe, fe_missing = build_frontend_workbook(fe_groups_data, today_str)
    wb_fe.save(FRONTEND_OUTPUT)
    total_pages = sum(len(ps) for _, ps in fe_groups_data)
    print(f"\nSaved => {FRONTEND_OUTPUT}")
    print(f"  Sheets          : {list(wb_fe.sheetnames)}")
    print(f"  Total pages     : {total_pages}")
    print(f"  Incomplete rows : {len(fe_missing)}  (see sheet: ⚠ Missing Frontend)")


if __name__ == "__main__":
    main()


# =============================================================================
# PUBLIC API — called from main.py via --excel flag
# =============================================================================

def generate_excel(
    docs_root: str,
    output_dir: str,
    project_name: str = None,
) -> tuple:
    """
    Generate backend and frontend Excel workbooks from docs in docs_root.

    Args:
        docs_root:    Path to docs/ folder (must contain backend/ and/or frontend/).
        output_dir:   Where to save the .xlsx files.
        project_name: Used for workbook titles and filenames. Defaults to output_dir basename.

    Returns:
        (backend_excel_path, frontend_excel_path) — either may be "" if not generated.
    """
    global PROJECT_NAME, BACKEND_PATH, FRONTEND_PATH, OUTPUT_DIR
    global BACKEND_OUTPUT, FRONTEND_OUTPUT

    PROJECT_NAME    = project_name or os.path.basename(os.path.abspath(output_dir))
    BACKEND_PATH    = os.path.join(docs_root, "backend")
    FRONTEND_PATH   = os.path.join(docs_root, "frontend")
    OUTPUT_DIR      = output_dir
    BACKEND_OUTPUT  = os.path.join(output_dir, "{}_backend.xlsx".format(PROJECT_NAME))
    FRONTEND_OUTPUT = os.path.join(output_dir, "{}_frontend.xlsx".format(PROJECT_NAME))

    os.makedirs(output_dir, exist_ok=True)
    main()

    be = BACKEND_OUTPUT  if os.path.exists(BACKEND_OUTPUT)  else ""
    fe = FRONTEND_OUTPUT if os.path.exists(FRONTEND_OUTPUT) else ""
    return be, fe
